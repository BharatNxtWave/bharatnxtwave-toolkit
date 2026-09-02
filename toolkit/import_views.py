import csv
import hashlib
import io
import zipfile
from pathlib import Path, PurePosixPath

from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from openpyxl import load_workbook

from accounts.activity import log_activity

from .import_forms import ToolkitImportUploadForm
from .models import ImportBatch
from .intelligence.ingestion import analyse_file
from .intelligence.staging import stage_analysis


MAX_PREVIEW_ROWS = 8
MAX_PREVIEW_COLUMNS = 30

MAX_XLSX_UNCOMPRESSED_BYTES = (
    250 * 1024 * 1024
)

MAX_XLSX_COMPRESSION_RATIO = 200


def can_import_toolkit(user):
    from accounts.portal_access import is_admin_user

    return is_admin_user(user)


def normalize_value(value):
    if value is None:
        return ""

    return str(value).strip()


# =========================================================
# FILE FINGERPRINT
# =========================================================

def calculate_sha256(uploaded_file):
    digest = hashlib.sha256()

    uploaded_file.seek(0)

    for chunk in uploaded_file.chunks():
        digest.update(chunk)

    uploaded_file.seek(0)

    return digest.hexdigest()


# =========================================================
# CLIENT-REPORTED FILE MODIFICATION DATE
#
# This comes from browser File.lastModified.
# It is useful operational metadata but is not treated as
# a cryptographically trusted timestamp.
# =========================================================

def parse_source_modified_at(value):
    if not value:
        return None

    parsed = parse_datetime(value)

    if parsed is None:
        return None

    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed)

    return parsed


# =========================================================
# XLSX SECURITY VALIDATION
# =========================================================

def validate_xlsx_structure(uploaded_file):
    uploaded_file.seek(0)

    if not zipfile.is_zipfile(uploaded_file):
        uploaded_file.seek(0)

        raise ValueError(
            "The Excel file is not a valid XLSX archive."
        )

    uploaded_file.seek(0)

    with zipfile.ZipFile(uploaded_file) as archive:
        infos = archive.infolist()

        names = {
            info.filename
            for info in infos
        }

        required_files = {
            "[Content_Types].xml",
            "xl/workbook.xml",
        }

        if not required_files.issubset(names):
            raise ValueError(
                "The file does not contain a valid XLSX workbook."
            )

        # XLSX only. Macro-enabled content is not accepted.
        if any(
            name.lower().endswith(
                "vbaproject.bin"
            )
            for name in names
        ):
            raise ValueError(
                "Macro-enabled Excel files are not permitted."
            )

        total_compressed = 0
        total_uncompressed = 0

        for info in infos:
            path = PurePosixPath(
                info.filename
            )

            if (
                info.filename.startswith("/")
                or ".." in path.parts
            ):
                raise ValueError(
                    "Unsafe XLSX archive structure detected."
                )

            if info.flag_bits & 0x1:
                raise ValueError(
                    "Encrypted XLSX archives are not supported."
                )

            total_compressed += (
                info.compress_size
            )

            total_uncompressed += (
                info.file_size
            )

        if (
            total_uncompressed
            > MAX_XLSX_UNCOMPRESSED_BYTES
        ):
            raise ValueError(
                "The expanded Excel workbook is too large."
            )

        if (
            total_compressed > 0
            and (
                total_uncompressed
                / total_compressed
            )
            > MAX_XLSX_COMPRESSION_RATIO
        ):
            raise ValueError(
                "Unsafe Excel compression ratio detected."
            )

    uploaded_file.seek(0)


# =========================================================
# CSV SECURITY VALIDATION + PREVIEW
# =========================================================

def preview_csv(uploaded_file):
    uploaded_file.seek(0)

    raw = uploaded_file.read()

    uploaded_file.seek(0)

    if b"\x00" in raw:
        raise ValueError(
            "Binary content detected in CSV file."
        )

    text = None

    for encoding in (
        "utf-8-sig",
        "utf-8",
        "cp1252",
    ):
        try:
            text = raw.decode(encoding)
            break

        except UnicodeDecodeError:
            continue

    if text is None:
        raise ValueError(
            "CSV encoding could not be read."
        )

    csv.field_size_limit(
        1_000_000
    )

    stream = io.StringIO(text)

    reader = csv.reader(stream)

    first_row = next(
        reader,
        None
    )

    if first_row is None:
        return {
            "sheets": [
                {
                    "name": "CSV",
                    "headers": [],
                    "preview_rows": [],
                    "total_rows": 0,
                    "total_columns": 0,
                }
            ]
        }

    headers = [
        normalize_value(value)
        for value
        in first_row[
            :MAX_PREVIEW_COLUMNS
        ]
    ]

    preview_rows = []

    total_rows = 0
    total_columns = len(first_row)

    for row in reader:
        total_rows += 1

        total_columns = max(
            total_columns,
            len(row)
        )

        if (
            len(preview_rows)
            < MAX_PREVIEW_ROWS
        ):
            preview_rows.append(
                [
                    normalize_value(value)
                    for value
                    in row[
                        :MAX_PREVIEW_COLUMNS
                    ]
                ]
            )

    return {
        "sheets": [
            {
                "name": "CSV",
                "headers": headers,
                "preview_rows": preview_rows,
                "total_rows": total_rows,
                "total_columns": total_columns,
            }
        ]
    }


# =========================================================
# XLSX PREVIEW
# =========================================================

def preview_excel(uploaded_file):
    validate_xlsx_structure(
        uploaded_file
    )

    uploaded_file.seek(0)

    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    sheets = []

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(
                values_only=True
            )

            first_row = next(
                rows,
                None
            )

            if first_row is None:
                sheets.append(
                    {
                        "name": worksheet.title,
                        "headers": [],
                        "preview_rows": [],
                        "total_rows": 0,
                        "total_columns": 0,
                    }
                )

                continue

            headers = [
                normalize_value(value)
                for value
                in first_row[
                    :MAX_PREVIEW_COLUMNS
                ]
            ]

            preview_rows = []

            total_data_rows = 0

            for row in rows:
                total_data_rows += 1

                if (
                    len(preview_rows)
                    < MAX_PREVIEW_ROWS
                ):
                    preview_rows.append(
                        [
                            normalize_value(value)
                            for value
                            in row[
                                :MAX_PREVIEW_COLUMNS
                            ]
                        ]
                    )

            sheets.append(
                {
                    "name": worksheet.title,
                    "headers": headers,
                    "preview_rows": preview_rows,
                    "total_rows": total_data_rows,
                    "total_columns": (
                        worksheet.max_column
                        or 0
                    ),
                }
            )

    finally:
        workbook.close()

        uploaded_file.seek(0)

    return {
        "sheets": sheets
    }




def _extraction_review_stats(batch):

    if batch is None:
        return None

    rows = batch.rows.all()

    candidates = rows.exclude(
        candidate_action="UNDECIDED"
    )

    return {
        "source_rows": rows.count(),

        "candidates": candidates.count(),

        "create": candidates.filter(
            candidate_action="CREATE"
        ).count(),

        "update": candidates.filter(
            candidate_action="UPDATE"
        ).count(),

        "merge_review": candidates.filter(
            candidate_action="MERGE_REVIEW"
        ).count(),

        "knowledge_rows": rows.filter(
            candidate_action="UNDECIDED"
        ).count(),
    }


# =========================================================
# IMPORT CENTER / SOURCE PREVIEW
# =========================================================

@user_passes_test(can_import_toolkit)
def import_center(request):
    preview = None
    import_error = None
    uploaded_name = None

    current_batch = None
    duplicate_batch = None
    analysis = None
    staging_result = None

    latest_review_batch = (
        ImportBatch.objects
        .filter(
            status__in=[
                "PREVIEWED",
                "VALIDATED",
            ],
            rows__isnull=False,
        )
        .exclude(
            metadata__duplicate_detected=True
        )
        .distinct()
        .order_by("-created_at")
        .first()
    )

    latest_review_stats = (
        _extraction_review_stats(
            latest_review_batch
        )
    )

    if request.method == "POST":
        form = ToolkitImportUploadForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():
            uploaded_file = (
                form.cleaned_data["file"]
            )

            uploaded_name = (
                uploaded_file.name
            )

            extension = (
                Path(uploaded_file.name)
                .suffix
                .lower()
            )

            source_type = (
                "CSV"
                if extension == ".csv"
                else "XLSX"
            )

            try:
                file_sha256 = calculate_sha256(
                    uploaded_file
                )

                # Identical content, even if renamed.
                duplicate_batch = (
                    ImportBatch.objects
                    .filter(
                        file_sha256=file_sha256
                    )
                    .order_by(
                        "-created_at"
                    )
                    .first()
                )

                # Compare current version against the most
                # recent source with the same filename.
                previous_same_name = (
                    ImportBatch.objects
                    .filter(
                        source_type=source_type,
                        source_name=uploaded_name,
                    )
                    .order_by(
                        "-created_at"
                    )
                    .first()
                )

                if extension == ".csv":
                    preview = preview_csv(
                        uploaded_file
                    )

                elif extension == ".xlsx":
                    preview = preview_excel(
                        uploaded_file
                    )

                else:
                    raise ValueError(
                        "Unsupported file format."
                    )

                analysis = analyse_file(
                    uploaded_file,
                    uploaded_name,
                )

                sheets = preview.get(
                    "sheets",
                    []
                )

                total_rows = sum(
                    sheet.get(
                        "total_rows",
                        0
                    )
                    for sheet in sheets
                )

                source_modified_at = (
                    parse_source_modified_at(
                        request.POST.get(
                            "source_modified_at"
                        )
                    )
                )

                metadata = {
                    "operation": "preview",
                    "preview_only": True,

                    "sheet_names": [
                        sheet.get(
                            "name",
                            ""
                        )
                        for sheet in sheets
                    ],

                    "sheets": [
                        {
                            "name": sheet.get(
                                "name",
                                ""
                            ),
                            "rows": sheet.get(
                                "total_rows",
                                0
                            ),
                            "columns": sheet.get(
                                "total_columns",
                                0
                            ),
                        }
                        for sheet in sheets
                    ],

                    "duplicate_detected": (
                        duplicate_batch
                        is not None
                    ),

                    "duplicate_of_id": (
                        duplicate_batch.pk
                        if duplicate_batch
                        else None
                    ),

                    "previous_same_name_id": (
                        previous_same_name.pk
                        if previous_same_name
                        else None
                    ),

                    "source_changed_since_previous": (
                        bool(
                            previous_same_name
                            and (
                                previous_same_name.file_sha256
                                != file_sha256
                            )
                        )
                    ),
                }

                current_batch = (
                    ImportBatch.objects.create(
                        source_type=source_type,

                        source_name=(
                            uploaded_name
                        ),

                        source_modified_at=(
                            source_modified_at
                        ),

                        file_sha256=file_sha256,

                        sheet_count=len(
                            sheets
                        ),

                        row_count=total_rows,

                        status="PREVIEWED",

                        imported_by=(
                            request.user
                        ),

                        metadata=metadata,
                    )
                )

                staging_result = stage_analysis(
                    current_batch,
                    analysis,
                    uploaded_file,
                    uploaded_name,
                )

                meaningful_rows = (
                    staging_result.get(
                        "source_rows"
                    )
                    if staging_result
                    else None
                )

                if meaningful_rows is not None:

                    metadata = dict(
                        current_batch.metadata
                        or {}
                    )

                    if (
                        current_batch.row_count
                        != meaningful_rows
                    ):
                        metadata[
                            "preview_dimension_row_count"
                        ] = current_batch.row_count

                    metadata[
                        "meaningful_source_row_count"
                    ] = meaningful_rows

                    current_batch.row_count = (
                        meaningful_rows
                    )

                    current_batch.metadata = (
                        metadata
                    )

                    current_batch.save(
                        update_fields=[
                            "row_count",
                            "metadata",
                        ]
                    )

                latest_review_batch = (
                    current_batch
                )

                latest_review_stats = (
                    _extraction_review_stats(
                        current_batch
                    )
                )

                # Do NOT put confidential filenames,
                # row contents or cell contents into the
                # generic ActivityLog.
                log_activity(
                    request,
                    "IMPORT",
                    "Previewed toolkit source file.",
                    target_type=(
                        "import_batch"
                    ),
                    target_id=(
                        current_batch.pk
                    ),
                    metadata={
                        "operation": (
                            "preview"
                        ),
                        "batch_id": (
                            current_batch.pk
                        ),
                        "source_type": (
                            source_type
                        ),
                        "sheet_count": (
                            len(sheets)
                        ),
                        "row_count": (
                            total_rows
                        ),
                        "duplicate_detected": (
                            duplicate_batch
                            is not None
                        ),
                    },
                )

            except Exception as exc:
                import_error = str(
                    exc
                ) or (
                    "The source file could "
                    "not be read."
                )

    else:
        form = ToolkitImportUploadForm()

    return render(
        request,
        "toolkit/admin/import_center.html",
        {
            "form": form,
            "preview": preview,
            "import_error": import_error,
            "uploaded_name": uploaded_name,
            "current_batch": current_batch,
            "duplicate_batch": duplicate_batch,
            "analysis": analysis,
            "staging_result": staging_result,
            "latest_review_batch": latest_review_batch,
            "latest_review_stats": latest_review_stats,
        }
    )


# =========================================================
# SOURCE / IMPORT HISTORY
# =========================================================


# BNW_IMPORT_STATUS_WORDING_V1
def _with_toolkit_change_status(queryset):
    """
    Add a display-only flag showing whether an imported
    source produced active Toolkit audit changes.

    The ImportBatch database status remains untouched.
    """

    from django.db.models import Exists, OuterRef
    from toolkit.models import ImportChange

    effective_changes = (
        ImportChange.objects
        .filter(
            import_batch_id=OuterRef("pk"),
            is_reversed=False,
        )
    )

    return queryset.annotate(
        has_toolkit_changes=Exists(
            effective_changes
        )
    )


@user_passes_test(can_import_toolkit)
def import_history(request):

    source_filter = request.GET.get(
        "source",
        "",
    ).strip()

    status_filter = request.GET.get(
        "status",
        "",
    ).strip()

    query = request.GET.get(
        "q",
        "",
    ).strip()

    batches = _with_toolkit_change_status(
        ImportBatch.objects
        .select_related(
            "imported_by"
        )
        .all()
    )

    valid_sources = {
        value
        for value, label
        in ImportBatch.SOURCE_CHOICES
    }

    valid_statuses = {
        value
        for value, label
        in ImportBatch.STATUS_CHOICES
    }

    if source_filter in valid_sources:

        batches = batches.filter(
            source_type=source_filter
        )

    if status_filter in valid_statuses:

        batches = batches.filter(
            status=status_filter
        )

    if query:

        batches = batches.filter(
            Q(
                source_name__icontains=query
            )
            | Q(
                file_sha256__icontains=query
            )
        )

    all_batches = (
        ImportBatch.objects.all()
    )

    latest_batch = (
        all_batches
        .order_by("-created_at")
        .first()
    )

    latest_review_batch = (
        all_batches
        .filter(
            status__in=[
                "PREVIEWED",
                "VALIDATED",
            ],
            rows__isnull=False,
        )
        .exclude(
            metadata__duplicate_detected=True
        )
        .distinct()
        .order_by("-created_at")
        .first()
    )

    latest_review_stats = (
        _extraction_review_stats(
            latest_review_batch
        )
    )

    duplicate_count = sum(
        1
        for metadata
        in all_batches.values_list(
            "metadata",
            flat=True,
        )
        if (
            isinstance(
                metadata,
                dict,
            )
            and metadata.get(
                "duplicate_detected"
            )
        )
    )

    unique_source_count = (
        all_batches
        .exclude(
            file_sha256=""
        )
        .values(
            "file_sha256"
        )
        .distinct()
        .count()
    )

    historical_rows_processed = (
        all_batches.aggregate(
            total=Sum("row_count")
        )["total"]
        or 0
    )

    context = {

        "batches":
            batches.order_by(
                "-created_at"
            )[:250],

        "source_filter":
            source_filter,

        "status_filter":
            status_filter,

        "query":
            query,

        "source_choices":
            ImportBatch.SOURCE_CHOICES,

        "status_choices":
            ImportBatch.STATUS_CHOICES,

        "total_batches":
            all_batches.count(),

        "duplicate_count":
            duplicate_count,

        "unique_source_count":
            unique_source_count,

        "latest_batch":
            latest_batch,

        "latest_source_rows":
            (
                latest_batch.row_count
                if latest_batch
                else 0
            ),

        "latest_review_batch":
            latest_review_batch,

        "latest_review_stats":
            latest_review_stats,

        # Audit metric only.
        # Not presented as current Toolkit data.
        "historical_rows_processed":
            historical_rows_processed,
    }

    return render(
        request,
        "toolkit/admin/import_history.html",
        context,
    )


# =========================================================
# SOURCE HISTORY DETAIL
# =========================================================

@user_passes_test(can_import_toolkit)
def import_history_detail(
    request,
    batch_id
):
    batch = get_object_or_404(
        _with_toolkit_change_status(
            ImportBatch.objects
            .select_related(
                "imported_by"
            )
        ),
        pk=batch_id
    )

    duplicate_of = None
    previous_same_name = None

    if isinstance(
        batch.metadata,
        dict
    ):
        duplicate_id = (
            batch.metadata.get(
                "duplicate_of_id"
            )
        )

        previous_id = (
            batch.metadata.get(
                "previous_same_name_id"
            )
        )

        if duplicate_id:
            duplicate_of = (
                ImportBatch.objects
                .filter(
                    pk=duplicate_id
                )
                .first()
            )

        if previous_id:
            previous_same_name = (
                ImportBatch.objects
                .filter(
                    pk=previous_id
                )
                .first()
            )

    return render(
        request,
        "toolkit/admin/import_history_detail.html",
        {
            "batch": batch,
            "duplicate_of": duplicate_of,
            "previous_same_name": (
                previous_same_name
            ),
        }
    )


# ============================================================
# SOURCE HISTORY RECORD MANAGEMENT
# ============================================================

from django.contrib import messages as _source_messages
from django.contrib.auth.decorators import login_required as _source_login_required
from django.http import HttpResponseForbidden as _SourceHttpResponseForbidden
from django.shortcuts import get_object_or_404 as _source_get_object_or_404
from django.shortcuts import redirect as _source_redirect
from django.views.decorators.http import require_POST as _source_require_POST


def _can_delete_source_record(user):
    from accounts.portal_access import is_admin_user

    return is_admin_user(user)


@_source_login_required
@_source_require_POST
def import_source_delete(
    request,
    batch_id,
):

    if not _can_delete_source_record(
        request.user
    ):

        return _SourceHttpResponseForbidden(
            "You do not have permission to delete source records."
        )


    batch = _source_get_object_or_404(
        ImportBatch,
        pk=batch_id,
    )


    # --------------------------------------------------------
    # Do not allow deletion of a source that has actually
    # been imported into the live Toolkit.
    # --------------------------------------------------------

    if str(
        batch.status
    ).upper() == "IMPORTED":

        _source_messages.error(
            request,
            (
                "This source has already been imported into the "
                "Toolkit and cannot be deleted from Source History."
            ),
        )

        return _source_redirect(
            f"/admin-center/import/history/{batch.pk}/"
        )


    event_id = batch.pk


    batch.delete()


    _source_messages.success(
        request,
        (
            f"Source Event #{event_id} was deleted "
            "from Source History."
        ),
    )


    return _source_redirect(
        "/admin-center/import/history/"
    )



# =========================================================
# INTELLIGENT EXTRACTION REVIEW
# =========================================================

from django.contrib import messages as _review_messages
from django.shortcuts import redirect as _review_redirect
from django.views.decorators.http import require_POST as _review_require_POST

from .models import (
    Category as _ReviewCategory,
    ImportRow as _ReviewImportRow,
    Service as _ReviewService,
)


@user_passes_test(can_import_toolkit)
def import_extraction_review(
    request,
    batch_id,
):

    batch = get_object_or_404(
        ImportBatch.objects,
        pk=batch_id,
    )

    batch_metadata = (
        batch.metadata
        if isinstance(
            batch.metadata,
            dict,
        )
        else {}
    )

    if batch_metadata.get(
        "duplicate_detected"
    ):

        duplicate_of = None

        duplicate_of_id = (
            batch_metadata.get(
                "duplicate_of_id"
            )
        )

        if duplicate_of_id:

            duplicate_of = (
                ImportBatch.objects
                .filter(
                    pk=duplicate_of_id
                )
                .first()
            )

        recognized_candidates = (
            batch.rows
            .exclude(
                candidate_action="UNDECIDED"
            )
            .count()
        )

        return render(
            request,
            "toolkit/admin/extraction_duplicate.html",
            {
                "batch":
                    batch,

                "duplicate_of":
                    duplicate_of,

                "recognized_candidates":
                    recognized_candidates,

                "meaningful_rows":
                    batch.rows.count(),
            },
        )

    all_rows = (
        batch.rows
        .select_related(
            "matched_service",
            "matched_service__category",
            "matched_service__domain",
        )
        .order_by(
            "sheet_name",
            "source_row_number",
        )
    )

    candidates = list(
        all_rows.exclude(
            candidate_action="UNDECIDED"
        )
    )

    categories = list(
        _ReviewCategory.objects
        .select_related(
            "domain"
        )
        .order_by(
            "domain__name",
            "name",
        )
    )

    category_slug_map = {}

    for category in categories:

        category_slug_map.setdefault(
            category.slug,
            category.pk,
        )

    # BNW_CHANGED_INFORMATION_REVIEW_UI_V1
    from toolkit.intelligence.final_import import (
        classify_candidate,
    )

    approved = 0
    skipped = 0

    for row in candidates:

        raw = (
            row.raw_data
            if isinstance(
                row.raw_data,
                dict,
            )
            else {}
        )

        candidate = (
            raw.get("candidate", {})
            if isinstance(
                raw.get(
                    "candidate",
                    {},
                ),
                dict,
            )
            else {}
        )

        proposal = (
            candidate.get(
                "proposal",
                {},
            )
            if isinstance(
                candidate.get(
                    "proposal",
                    {},
                ),
                dict,
            )
            else {}
        )

        review = (
            raw.get(
                "review",
                {},
            )
            if isinstance(
                raw.get(
                    "review",
                    {},
                ),
                dict,
            )
            else {}
        )

        row.review_meta = review

        row.review_decision = (
            review.get(
                "decision"
            )
            or "PENDING"
        )

        if (
            row.review_decision
            == "APPROVED"
        ):
            approved += 1

        if (
            row.candidate_action
            == "SKIP"
        ):
            skipped += 1

        proposed_slug = str(
            proposal.get(
                "category_slug",
                "",
            )
            or ""
        )

        row.selected_category_id = (
            review.get(
                "category_id"
            )
            or category_slug_map.get(
                proposed_slug
            )
        )

        row.display_category_name = (
            review.get(
                "category_name"
            )
            or proposed_slug
            or "Needs Review"
        )

        row.display_service_kind = (
            review.get(
                "service_kind"
            )
            or proposal.get(
                "service_kind"
            )
            or "OTHER"
        )

        try:

            internal_signal = float(
                proposal.get(
                    "confidence",
                    0,
                )
                or 0
            )

        except (
            TypeError,
            ValueError,
        ):

            internal_signal = 0

        # Internal only.
        # Numeric value is never displayed.
        row.is_clear_suggestion = (
            internal_signal >= 0.8
        )

        if row.candidate_action == "SKIP":

            delta = {
                "status": "SKIPPED",
                "reason": "Item excluded by the administrator.",
                "differences": [],
            }

        elif row.candidate_action == "CREATE":

            delta = {
                "status": "NEW_SERVICE",
                "reason": "No existing Service match.",
                "differences": [],
            }

        else:

            try:
                delta = classify_candidate(
                    candidate
                )

            except Exception:

                # Fail safely. If comparison cannot be completed,
                # the row must require attention.
                delta = {
                    "status": "CONFLICT",
                    "reason": (
                        "The incoming information could not be "
                        "safely compared."
                    ),
                    "differences": [],
                }

        row.delta_status = str(
            delta.get(
                "status",
                "CONFLICT",
            )
            or "CONFLICT"
        )

        row.delta_reason = str(
            delta.get(
                "reason",
                "",
            )
            or ""
        )

        delta_differences = delta.get(
            "differences",
            [],
        )

        row.delta_differences = (
            delta_differences
            if isinstance(
                delta_differences,
                list,
            )
            else []
        )

        row.is_changed_information = (
            row.candidate_action != "SKIP"
            and row.delta_status
            == "CHANGED_INFORMATION"
        )

        if row.candidate_action == "SKIP":

            row.needs_attention = False

        elif row.candidate_action == "CREATE":

            row.needs_attention = (
                row.delta_status
                != "NEW_SERVICE"
            )

        elif row.candidate_action == "UPDATE":

            row.needs_attention = (
                row.delta_status
                not in {
                    "NO_CHANGE",
                    "SAFE_ADDITION",
                }
            )

        else:

            row.needs_attention = True

        row.can_approve_import = (
            row.candidate_action
            != "SKIP"
            and not row.needs_attention
        )

    stats = {

        "source_rows":
            all_rows.count(),

        "candidates":
            len(candidates),

        "create":
            sum(
                row.candidate_action
                == "CREATE"
                for row in candidates
            ),

        "update":
            sum(
                row.candidate_action
                == "UPDATE"
                for row in candidates
            ),

        "safe_update":
            sum(
                row.candidate_action
                == "UPDATE"
                and not row.needs_attention
                for row in candidates
            ),

        "merge_review":
            sum(
                row.candidate_action
                == "MERGE_REVIEW"
                for row in candidates
            ),

        "needs_attention":
            sum(
                row.needs_attention
                for row in candidates
            ),

        "blocked_replacements":
            sum(
                row.is_changed_information
                for row in candidates
            ),

        "approved":
            sum(
                row.review_decision
                == "APPROVED"
                and not row.needs_attention
                for row in candidates
            ),

        "skipped":
            skipped,

        "knowledge_rows":
            all_rows.filter(
                candidate_action="UNDECIDED"
            ).count(),
    }

    return render(
        request,
        "toolkit/admin/extraction_review.html",
        {
            "batch":
                batch,

            "rows":
                candidates,

            "stats":
                stats,

            "categories":
                categories,

            "service_kind_choices":
                _ReviewService
                .SERVICE_KIND_CHOICES,
        },
    )


@user_passes_test(can_import_toolkit)
@_review_require_POST
def import_extraction_row_decision(
    request,
    batch_id,
    row_id,
):

    batch = get_object_or_404(
        ImportBatch.objects,
        pk=batch_id,
    )

    row = get_object_or_404(
        _ReviewImportRow.objects,
        pk=row_id,
        import_batch=batch,
    )

    if batch.status not in {
        "PREVIEWED",
        "VALIDATED",
    }:

        _review_messages.error(
            request,
            (
                "This source is no longer "
                "available for review."
            ),
        )

        return _review_redirect(
            "toolkit:import_extraction_review",
            batch_id=batch.pk,
        )

    decision = str(
        request.POST.get(
            "decision",
            "",
        )
    ).strip()

    raw = (
        dict(row.raw_data)
        if isinstance(
            row.raw_data,
            dict,
        )
        else {}
    )

    review = (
        dict(
            raw.get(
                "review",
                {},
            )
        )
        if isinstance(
            raw.get(
                "review",
                {},
            ),
            dict,
        )
        else {}
    )

    candidate = (
        raw.get(
            "candidate",
            {},
        )
        if isinstance(
            raw.get(
                "candidate",
                {},
            ),
            dict,
        )
        else {}
    )

    match = (
        candidate.get(
            "match",
            {},
        )
        if isinstance(
            candidate.get(
                "match",
                {},
            ),
            dict,
        )
        else {}
    )

    now = timezone.now().isoformat()


    # --------------------------------------------------------
    # SAVE CLASSIFICATION
    # --------------------------------------------------------

    if decision == "save_mapping":

        category_id = (
            request.POST.get(
                "category_id"
            )
        )

        service_kind = str(
            request.POST.get(
                "service_kind",
                "",
            )
        ).strip()

        category = (
            _ReviewCategory.objects
            .select_related(
                "domain"
            )
            .filter(
                pk=category_id
            )
            .first()
        )

        valid_kinds = {
            value
            for value, label
            in _ReviewService
            .SERVICE_KIND_CHOICES
        }

        if category is None:

            _review_messages.error(
                request,
                "Please choose a valid category.",
            )

        elif (
            service_kind
            not in valid_kinds
        ):

            _review_messages.error(
                request,
                (
                    "Please choose a valid "
                    "Service type."
                ),
            )

        else:

            review.update(
                {
                    "category_id":
                        category.pk,

                    "category_slug":
                        category.slug,

                    "category_name":
                        category.name,

                    "domain_slug":
                        category.domain.slug,

                    "service_kind":
                        service_kind,

                    "classification_edited":
                        True,

                    "classification_edited_at":
                        now,

                    "classification_edited_by_id":
                        request.user.pk,
                }
            )

            _review_messages.success(
                request,
                "Classification saved.",
            )


    # --------------------------------------------------------
    # APPROVE
    # --------------------------------------------------------

    elif decision == "approve":

        from toolkit.intelligence.final_import import (
            classify_candidate as
            _classify_candidate_for_approval,
        )

        try:
            approval_delta_status = str(
                _classify_candidate_for_approval(
                    candidate
                ).get(
                    "status",
                    "",
                )
                or ""
            )

        except Exception:
            approval_delta_status = ""

        if (
            approval_delta_status
            == "CHANGED_INFORMATION"
        ):

            messages.error(
                request,
                (
                    "Approval blocked: this file replaces "
                    "information already stored in the Toolkit. "
                    "Choose 'Do not import this item', or update "
                    "the Service manually after verification."
                ),
            )

            return redirect(
                "toolkit:import_extraction_review",
                batch_id=batch.pk,
            )

        if (
            row.candidate_action
            == "MERGE_REVIEW"
        ):

            if row.matched_service_id:

                row.candidate_action = (
                    "UPDATE"
                )

            else:

                _review_messages.error(
                    request,
                    (
                        "This item still needs "
                        "an existing Service match."
                    ),
                )

                return _review_redirect(
                    (
                        f"/admin-center/import/"
                        f"{batch.pk}/review/"
                        f"#row-{row.pk}"
                    )
                )

        if (
            row.candidate_action
            not in {
                "CREATE",
                "UPDATE",
            }
        ):

            _review_messages.error(
                request,
                (
                    "This item cannot be "
                    "approved yet."
                ),
            )

            return _review_redirect(
                (
                    f"/admin-center/import/"
                    f"{batch.pk}/review/"
                    f"#row-{row.pk}"
                )
            )

        review.update(
            {
                "decision":
                    "APPROVED",

                "reviewed_at":
                    now,

                "reviewed_by_id":
                    request.user.pk,
            }
        )

        row.validation_status = (
            "VALID"
        )

        _review_messages.success(
            request,
            (
                "Approved for final import. "
                "The live Service Library "
                "has not been changed yet."
            ),
        )


    # --------------------------------------------------------
    # SKIP
    # --------------------------------------------------------

    elif decision == "skip":

        if (
            row.candidate_action
            != "SKIP"
        ):

            review[
                "previous_action"
            ] = row.candidate_action

        review.update(
            {
                "decision":
                    "SKIPPED",

                "reviewed_at":
                    now,

                "reviewed_by_id":
                    request.user.pk,
            }
        )

        row.candidate_action = (
            "SKIP"
        )

        row.validation_status = (
            "PROCESSED"
        )

        _review_messages.success(
            request,
            (
                "Incoming change skipped "
                "for this review."
            ),
        )


    # --------------------------------------------------------
    # RESTORE
    # --------------------------------------------------------

    elif decision == "restore":

        original_action = (
            review.get(
                "previous_action"
            )
            or match.get(
                "action"
            )
            or "UNDECIDED"
        )

        valid_actions = {
            "UNDECIDED",
            "CREATE",
            "UPDATE",
            "MERGE_REVIEW",
            "INVALID",
        }

        if (
            original_action
            not in valid_actions
        ):

            original_action = (
                "UNDECIDED"
            )

        row.candidate_action = (
            original_action
        )

        if (
            original_action
            == "MERGE_REVIEW"
        ):

            row.validation_status = (
                "WARNING"
            )

        elif (
            original_action
            in {
                "CREATE",
                "UPDATE",
            }
        ):

            row.validation_status = (
                "VALID"
            )

        else:

            row.validation_status = (
                "PENDING"
            )

        review.update(
            {
                "decision":
                    "PENDING",

                "restored_at":
                    now,

                "restored_by_id":
                    request.user.pk,
            }
        )

        _review_messages.success(
            request,
            "Item restored to review.",
        )


    else:

        _review_messages.error(
            request,
            "Unknown review action.",
        )

        return _review_redirect(
            (
                f"/admin-center/import/"
                f"{batch.pk}/review/"
                f"#row-{row.pk}"
            )
        )


    raw["review"] = review

    row.raw_data = raw

    row.save(
        update_fields=[
            "raw_data",
            "candidate_action",
            "validation_status",
        ]
    )

    return _review_redirect(
        (
            f"/admin-center/import/"
            f"{batch.pk}/review/"
            f"#row-{row.pk}"
        )
    )


# ============================================================
# ADMIN_IMPORT_UX_V2
# LOW-CLICK SAFE APPROVAL
# ============================================================

@user_passes_test(can_import_toolkit)
def import_bulk_safe_approve(
    request,
    batch_id,
):
    """
    Approve all candidates that the existing pipeline has
    already classified as safe and VALID.

    This changes ImportRow review metadata only.
    It does NOT write to live Services.

    MERGE_REVIEW, WARNING, INVALID and ambiguous rows remain
    manual.
    """

    if request.method != "POST":

        return _review_redirect(
            "toolkit:import_extraction_review",
            batch_id=batch_id,
        )


    batch = get_object_or_404(
        ImportBatch.objects,
        pk=batch_id,
    )


    if batch.status not in {
        "PREVIEWED",
        "VALIDATED",
    }:

        _review_messages.error(
            request,
            (
                "This source is no longer "
                "available for approval."
            ),
        )

        return _review_redirect(
            "toolkit:import_extraction_review",
            batch_id=batch.pk,
        )


    from django.db import (
        transaction as _ux_transaction,
    )


    approved = 0
    already_approved = 0
    not_ready = 0


    with _ux_transaction.atomic():

        rows = (
            _ReviewImportRow.objects
            .select_for_update()
            .filter(
                import_batch=batch,
                candidate_action__in=[
                    "CREATE",
                    "UPDATE",
                ],
                validation_status="VALID",
            )
            .order_by(
                "sheet_name",
                "source_row_number",
            )
        )


        for row in rows:

            raw = (
                dict(row.raw_data)
                if isinstance(
                    row.raw_data,
                    dict,
                )
                else {}
            )


            review = (
                dict(
                    raw.get(
                        "review",
                        {},
                    )
                )
                if isinstance(
                    raw.get(
                        "review",
                        {},
                    ),
                    dict,
                )
                else {}
            )


            current_decision = str(
                review.get(
                    "decision",
                    "",
                )
                or ""
            ).upper()


            if current_decision == "APPROVED":

                already_approved += 1
                continue


            if current_decision == "SKIPPED":

                continue


            # UPDATE must have a real existing Service.
            if (
                row.candidate_action == "UPDATE"
                and not row.matched_service_id
            ):

                not_ready += 1
                continue


            review.update(
                {
                    "decision":
                        "APPROVED",

                    "reviewed_at":
                        timezone.now()
                        .isoformat(),

                    "reviewed_by_id":
                        request.user.pk,

                    "approved_by":
                        "ADMIN_IMPORT_UX_V2",

                    "bulk_safe_approval":
                        True,
                }
            )


            raw["review"] = review

            row.raw_data = raw

            row.validation_status = (
                "VALID"
            )


            row.save(
                update_fields=[
                    "raw_data",
                    "validation_status",
                ]
            )


            approved += 1


    if approved:

        _review_messages.success(
            request,
            (
                f"{approved} ready item"
                f"{'s' if approved != 1 else ''} "
                "approved together. "
                "No live Toolkit data has "
                "been changed yet."
            ),
        )

    elif already_approved:

        _review_messages.info(
            request,
            (
                "All ready items are already "
                "approved."
            ),
        )

    else:

        _review_messages.info(
            request,
            (
                "There are no automatic approvals "
                "waiting."
            ),
        )


    if not_ready:

        _review_messages.warning(
            request,
            (
                f"{not_ready} item"
                f"{'s' if not_ready != 1 else ''} "
                "still need manual review."
            ),
        )


    return _review_redirect(
        "toolkit:import_extraction_review",
        batch_id=batch.pk,
    )

# BNW_STANDARD_IMPORT_FINALIZE_V1
#
# Standard uploaded XLSX/CSV batches must not be sent through
# reconciliation_finalize. Reconciliation remains a separate,
# restricted workflow.

from django.contrib import messages as _standard_import_messages
from django.contrib.auth.decorators import (
    user_passes_test as _standard_import_user_passes_test,
)
from django.db import transaction as _standard_import_transaction
from django.shortcuts import (
    get_object_or_404 as _standard_import_get_object_or_404,
    redirect as _standard_import_redirect,
)
from django.utils import timezone as _standard_import_timezone
from django.views.decorators.http import (
    require_POST as _standard_import_require_POST,
)

from .models import (
    ImportBatch as _StandardImportBatch,
    ImportRow as _StandardImportRow,
)
from .intelligence.final_import import (
    apply_batch as _standard_apply_batch,
)


def _standard_approve_ready_rows(
    batch,
    user,
):
    """
    Approve only rows already classified as VALID CREATE/UPDATE
    candidates. Ambiguous, invalid and skipped rows remain untouched.
    """

    approved = 0

    rows = (
        _StandardImportRow.objects
        .select_for_update()
        .filter(
            import_batch=batch,
            candidate_action__in=[
                "CREATE",
                "UPDATE",
            ],
            validation_status="VALID",
        )
        .order_by(
            "sheet_name",
            "source_row_number",
        )
    )

    for row in rows:

        raw = (
            dict(row.raw_data)
            if isinstance(row.raw_data, dict)
            else {}
        )

        review = (
            dict(raw.get("review", {}))
            if isinstance(
                raw.get("review", {}),
                dict,
            )
            else {}
        )

        decision = str(
            review.get("decision", "")
            or ""
        ).upper()

        if decision == "SKIPPED":
            continue

        if (
            row.candidate_action == "UPDATE"
            and not row.matched_service_id
        ):
            continue

        if decision != "APPROVED":

            review.update(
                {
                    "decision": "APPROVED",
                    "reviewed_at": (
                        _standard_import_timezone
                        .now()
                        .isoformat()
                    ),
                    "reviewed_by_id": user.pk,
                    "approved_by": (
                        "STANDARD_IMPORT_FINALIZE_V1"
                    ),
                    "bulk_safe_approval": True,
                }
            )

            raw["review"] = review
            row.raw_data = raw
            row.validation_status = "VALID"

            row.save(
                update_fields=[
                    "raw_data",
                    "validation_status",
                ]
            )

            approved += 1

    return approved


@_standard_import_user_passes_test(
    can_import_toolkit
)
@_standard_import_require_POST
def import_finalize(
    request,
    batch_id,
):
    """
    One-click final import for ordinary XLSX/CSV uploads.

    Safe rows are approved and imported within one outer transaction.
    apply_batch still performs duplicate, pending-review, conflict,
    backup and integrity checks.
    """

    batch = _standard_import_get_object_or_404(
        _StandardImportBatch,
        pk=batch_id,
    )

    metadata = (
        batch.metadata
        if isinstance(batch.metadata, dict)
        else {}
    )

    if metadata.get("reconciliation_mode"):

        _standard_import_messages.error(
            request,
            (
                "This source belongs to the "
                "reconciliation workflow."
            ),
        )

        return _standard_import_redirect(
            "toolkit:reconciliation_finalize",
            batch_id=batch.pk,
        )

    if batch.status not in {
        "PREVIEWED",
        "VALIDATED",
    }:

        _standard_import_messages.error(
            request,
            (
                "This source is no longer available "
                "for final import."
            ),
        )

        return _standard_import_redirect(
            "toolkit:import_extraction_review",
            batch_id=batch.pk,
        )

    action = str(
        request.POST.get(
            "action",
            "",
        )
    ).strip()

    if action != "final_import":

        _standard_import_messages.error(
            request,
            "Invalid final import request.",
        )

        return _standard_import_redirect(
            "toolkit:import_extraction_review",
            batch_id=batch.pk,
        )

    try:

        with _standard_import_transaction.atomic():

            locked_batch = (
                _StandardImportBatch.objects
                .select_for_update()
                .get(pk=batch.pk)
            )

            approved = (
                _standard_approve_ready_rows(
                    locked_batch,
                    request.user,
                )
            )

            result = _standard_apply_batch(
                locked_batch.pk,
                user=request.user,
                reconcile=False,
            )

    except Exception as exc:

        _standard_import_messages.error(
            request,
            (
                f"Import failed: {exc}. "
                "Nothing was imported."
            ),
        )

        return _standard_import_redirect(
            "toolkit:import_extraction_review",
            batch_id=batch.pk,
        )

    created_services = int(
        result.get(
            "created_services",
            0,
        )
        or 0
    )

    changes_created = int(
        result.get(
            "changes_created",
            0,
        )
        or 0
    )

    skipped_count = (
        batch.rows
        .filter(
            candidate_action="SKIP"
        )
        .count()
    )

    if (
        skipped_count > 0
        and created_services == 0
        and changes_created == 0
        and approved == 0
    ):

        _standard_import_messages.success(
            request,
            (
                "Review completed. "
                f"{skipped_count} skipped item"
                f"{'s were' if skipped_count != 1 else ' was'} "
                "excluded from this import. "
                "No Toolkit data was changed."
            ),
        )

    else:

        _standard_import_messages.success(
            request,
            (
                "Toolkit import completed. "
                f"{created_services} new Service"
                f"{'s' if created_services != 1 else ''} "
                "created and "
                f"{changes_created} audited change"
                f"{'s' if changes_created != 1 else ''} "
                "recorded. "
                f"{approved} ready item"
                f"{'s' if approved != 1 else ''} "
                f"{'were' if approved != 1 else 'was'} "
                "approved during import."
            ),
        )

    return _standard_import_redirect(
        "toolkit:import_history_detail",
        batch_id=batch.pk,
    )
