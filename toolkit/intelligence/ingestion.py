import csv
import hashlib
import io
import re

from collections import Counter
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

from toolkit.models import Category, Service


ALIASES = {
    "scheme_name": (
        "scheme",
        "scheme name",
        "service",
        "service name",
        "program",
        "programme",
    ),
    "benefits": (
        "benefit",
        "benefits",
        "assistance",
        "support",
    ),
    "eligibility": (
        "eligibility",
        "eligible",
        "eligibility criteria",
    ),
    "deadline": (
        "deadline",
        "last date",
        "application deadline",
        "closing date",
    ),
    "focus_sectors": (
        "sector",
        "sectors",
        "industry",
        "industries",
    ),
    "funding_organisation": (
        "funding organisation",
        "funding organization",
        "organisation",
        "organization",
    ),
    "scheme_type": (
        "scheme type",
    ),
    "applicable_for": (
        "applicable for",
        "who can apply",
    ),
    "portal_link": (
        "portal link",
        "application link",
        "apply link",
        "website",
        "url",
    ),
    "flyer": (
        "flyer",
        "brochure",
    ),
    "additional_info": (
        "additional info",
        "additional information",
        "reference",
    ),
    "government_charge": (
        "government fee",
        "government fees",
        "govt fee",
        "govt fees",
    ),
    "minimum_charge": (
        "minimum charge",
        "minimum charges",
        "consultancy fee",
        "our fee",
    ),
    "vendor_cost": (
        "vendor cost",
        "vendor costs",
        "govt fees vendor cost",
    ),
    "bdm_deduction": (
        "bdm deduction",
        "bdm deductions",
    ),
    "commercial_remark": (
        "remark",
        "remarks",
    ),
}


SECTIONS = (
    (
        "ELIGIBILITY",
        (
            "eligibility",
            "eligible",
            "minimum requirement",
            "who can apply",
        ),
    ),
    (
        "DOCUMENTS",
        (
            "documents required",
            "required documents",
            "checklist",
        ),
    ),
    (
        "PROCESS",
        (
            "how to",
            "process",
            "steps to",
            "procedure",
            "incorporate",
        ),
    ),
    (
        "BENEFITS",
        (
            "benefit",
            "advantages",
        ),
    ),
    (
        "FUNDING",
        (
            "funding",
            "fund scheme",
            "financial assistance",
        ),
    ),
    (
        "TIMELINE",
        (
            "timeline",
            "working days",
            "processing time",
        ),
    ),
    (
        "SCOPE",
        (
            "scope of work",
            "descriptive scope",
        ),
    ),
    (
        "COMMERCIAL",
        (
            "commercial",
            "amount to be charge",
            "charges",
            "consultancy fee",
        ),
    ),
    (
        "NOTES",
        (
            "please note",
            "note:",
        ),
    ),
)


CATEGORY_RULES = (
    (
        "startup-and-msme-recognition",
        "REGISTRATION",
        (
            "startup india",
            "dpiit recognition",
            "start-up india",
        ),
    ),
    (
        "company-incorporation",
        "REGISTRATION",
        (
            "private limited",
            "pvt ltd",
            "llp incorporation",
            "llp registration",
            "company incorporation",
            "opc",
        ),
    ),
    (
        "gst-and-tax-compliance",
        "COMPLIANCE",
        (
            "gst",
            "income tax",
            "80-iac",
            "80iac",
            "tax exemption",
        ),
    ),
    (
        "government-grants",
        "GRANT",
        (
            "grant",
        ),
    ),
    (
        "credit-and-guarantee-schemes",
        "GOVT_SCHEME",
        (
            "credit guarantee",
            "loan scheme",
        ),
    ),
    (
        "funding-programs",
        "GOVT_SCHEME",
        (
            "seed fund",
            "funding program",
            "funding programme",
        ),
    ),
    (
        "equity-funding",
        "EQUITY",
        (
            "equity",
            "venture fund",
            "vc fund",
            "angel fund",
        ),
    ),
    (
        "debt-funding",
        "DEBT",
        (
            "debt funding",
        ),
    ),
    (
        "licenses",
        "REGISTRATION",
        (
            "license",
            "licence",
        ),
    ),
    (
        "certifications",
        "CERTIFICATION",
        (
            "certificate",
            "certification",
            "iso",
            "zed",
            "nsic",
        ),
    ),
    (
        "trademark-and-intellectual-property",
        "LEGAL",
        (
            "trademark",
            "patent",
            "ipr",
        ),
    ),
    (
        "website-development",
        "DIGITAL",
        (
            "website",
            "web development",
        ),
    ),
    (
        "digital-marketing",
        "DIGITAL",
        (
            "digital marketing",
        ),
    ),
    (
        "seo",
        "DIGITAL",
        (
            "seo",
            "search engine optimization",
        ),
    ),
    (
        "social-media",
        "DIGITAL",
        (
            "social media",
        ),
    ),
)


def norm(value):
    return re.sub(
        r"[^a-z0-9]+",
        " ",
        str(value or "").lower(),
    ).strip()


def safe(value):
    if value is None:
        return ""

    if isinstance(
        value,
        (
            datetime,
            date,
        ),
    ):
        return value.isoformat()

    if isinstance(
        value,
        (
            str,
            int,
            float,
            bool,
        ),
    ):
        return value

    return str(value)


def sha256(file_obj):
    file_obj.seek(0)

    digest = hashlib.sha256()

    for chunk in iter(
        lambda: file_obj.read(
            1024 * 1024
        ),
        b"",
    ):
        digest.update(chunk)

    file_obj.seek(0)

    return digest.hexdigest()


def field_for_header(value):
    text = norm(value)

    for field, aliases in ALIASES.items():
        for alias in aliases:
            if (
                text == alias
                or (
                    len(alias) >= 7
                    and alias in text
                )
            ):
                return field

    return None


def detect_header(rows):
    best = None

    for row_no, values, _ in rows[:30]:

        mapped = {}

        nonempty = sum(
            bool(norm(value))
            for value in values
        )

        for index, value in enumerate(
            values
        ):
            field = field_for_header(
                value
            )

            if (
                field
                and field not in mapped
            ):
                mapped[field] = index

        score = (
            len(mapped) * 10
            + min(nonempty, 10)
        )

        if (
            mapped
            and (
                best is None
                or score > best["score"]
            )
        ):
            best = {
                "row": row_no,
                "mapping": mapped,
                "nonempty": nonempty,
                "score": score,
            }

    return best


def sheet_kind(rows, header):

    values = [
        str(value)
        for _, row, _ in rows
        for value in row
        if str(value or "").strip()
    ]

    joined = " ".join(
        norm(value)
        for value in values[:1000]
    )

    max_width = max(
        (
            len(row)
            for _, row, _ in rows
        ),
        default=0,
    )

    mapping = (
        header["mapping"]
        if header
        else {}
    )

    # --------------------------------------------------------
    # DATA SHAPE
    # --------------------------------------------------------

    data_rows = []

    if header:

        for row_no, row, links in rows:

            if row_no > header["row"]:
                data_rows.append(row)

    else:

        data_rows = [
            row
            for _, row, _ in rows
        ]

    meaningful_data_rows = [
        row
        for row in data_rows
        if any(
            str(value or "").strip()
            for value in row
        )
    ]

    multi_cell_rows = sum(
        sum(
            bool(str(value or "").strip())
            for value in row
        ) >= 2
        for row in meaningful_data_rows
    )

    pair_ratio = (
        multi_cell_rows
        / len(meaningful_data_rows)
        if meaningful_data_rows
        else 0
    )

    # --------------------------------------------------------
    # COMMERCIAL SERVICE TABLE
    #
    # Do NOT classify a narrative knowledge sheet as a
    # commercial catalogue simply because the text contains
    # words such as consultancy fee.
    #
    # We require actual commercial COLUMNS with repeated data.
    # --------------------------------------------------------

    commercial_fields = {
        "vendor_cost",
        "bdm_deduction",
        "minimum_charge",
        "government_charge",
        "commercial_remark",
    }

    actual_commercial_fields = (
        commercial_fields
        & set(mapping)
    )

    if (
        header
        and "scheme_name" in mapping
        and actual_commercial_fields
    ):

        title_index = mapping[
            "scheme_name"
        ]

        commercial_indexes = [
            mapping[field]
            for field
            in actual_commercial_fields
            if field in mapping
        ]

        titled = 0
        commercial_rows = 0

        for row in meaningful_data_rows:

            title = (
                row[title_index]
                if title_index < len(row)
                else ""
            )

            if not str(
                title or ""
            ).strip():
                continue

            titled += 1

            has_commercial = any(
                index < len(row)
                and str(
                    row[index] or ""
                ).strip()
                for index
                in commercial_indexes
            )

            if has_commercial:
                commercial_rows += 1

        commercial_ratio = (
            commercial_rows / titled
            if titled
            else 0
        )

        if (
            commercial_rows >= 5
            and commercial_ratio >= 0.35
        ):
            return "COMMERCIAL_SERVICE_TABLE"

    # --------------------------------------------------------
    # COMMUNICATION
    # --------------------------------------------------------

    communication_signals = sum(
        marker in joined
        for marker in (
            "subject",
            "dear ",
            "regards",
            "thank you",
            "email",
            "mail",
            "attached",
        )
    )

    knowledge_signals = sum(
        marker in joined
        for marker in (
            "eligibility",
            "eligibility criteria",
            "documents required",
            "required documents",
            "scope of work",
            "benefits",
            "objectives",
            "timeline",
            "how to",
            "minimum requirements",
            "financial assistance",
            "incorporation",
            "registration process",
        )
    )

    if (
        max_width <= 2
        and communication_signals >= 3
        and knowledge_signals <= 1
    ):
        return "COMMUNICATION"

    # --------------------------------------------------------
    # HORIZONTAL COMPARISON MATRIX
    # --------------------------------------------------------

    if header:

        nonempty = header["nonempty"]
        recognised = len(mapping)

        unrecognised = max(
            0,
            nonempty - recognised,
        )

        if (
            "scheme_name" in mapping
            and nonempty >= 4
            and unrecognised >= 3
            and recognised <= 2
        ):
            return "COMPARISON_MATRIX"

    # --------------------------------------------------------
    # NORMAL STRUCTURED SERVICE TABLE
    #
    # IMPORTANT:
    # Check this BEFORE narrative detection.
    #
    # A real scheme table naturally contains columns such as
    # Benefits and Eligibility. Those words must not make the
    # whole table look like narrative knowledge.
    # --------------------------------------------------------

    if (
        header
        and "scheme_name" in mapping
        and len(mapping) >= 2
    ):
        return "STRUCTURED_TABLE"

    # --------------------------------------------------------
    # SIMPLE REFERENCE TABLE
    #
    # Example:
    # Sr No | Benefit
    # --------------------------------------------------------

    if (
        header
        and "scheme_name" not in mapping
        and header["nonempty"] <= 3
        and pair_ratio >= 0.60
    ):
        return "REFERENCE_TABLE"

    # --------------------------------------------------------
    # NARRATIVE KNOWLEDGE
    # --------------------------------------------------------

    long_ratio = (
        sum(
            len(value.strip()) >= 35
            for value in values
        )
        / len(values)
        if values
        else 0
    )

    if (
        knowledge_signals >= 2
        and (
            long_ratio >= 0.08
            or (
                max_width <= 3
                and pair_ratio < 0.60
            )
        )
    ):
        return "NARRATIVE_KNOWLEDGE"

    if header:
        return "REFERENCE_TABLE"

    return "UNKNOWN"


def section_type(text):

    normalized = norm(text)

    for section, markers in SECTIONS:

        if any(
            marker in normalized
            for marker in markers
        ):
            return section

    return None


def is_heading(text):

    raw = str(
        text or ""
    ).strip()

    normalized = norm(raw)

    return bool(
        raw
        and (
            (
                raw.startswith("*")
                and raw.endswith("*")
            )
            or (
                len(raw) <= 110
                and section_type(raw)
            )
            or (
                len(raw) <= 90
                and (
                    raw.endswith(":")
                    or normalized
                    in {
                        "benefits",
                        "commercials",
                        "timeline",
                    }
                )
            )
        )
    )


def propose(text, available):

    normalized = norm(text)

    for (
        category_slug,
        service_kind,
        markers,
    ) in CATEGORY_RULES:

        if (
            category_slug in available
            and any(
                marker in normalized
                for marker in markers
            )
        ):
            return {
                "category_slug":
                    category_slug,

                "service_kind":
                    service_kind,

                "confidence":
                    0.85,
            }

    fallback = (
        "other-business-services"
        if (
            "other-business-services"
            in available
        )
        else None
    )

    return {
        "category_slug":
            fallback,

        "service_kind":
            "OTHER",

        "confidence":
            0.35,
    }



def _canonical_service_title(value):

    text = norm(
        value
    )

    if not text:

        return ""

    replacements = (
        (
            r"\bstart\s+up\b",
            "startup",
        ),
        (
            r"\b80\s+iac\b",
            "80iac",
        ),
        (
            r"\bpvt\s+ltd\b",
            "private limited",
        ),
        (
            r"\bpvt\s+limited\b",
            "private limited",
        ),
        (
            r"\bpvt\b",
            "private",
        ),
        (
            r"\bltd\b",
            "limited",
        ),
        (
            r"\bllp\b",
            "limited liability partnership",
        ),
        (
            r"\bincorporation\b",
            "registration",
        ),
        (
            r"\bincorporate\b",
            "registration",
        ),
        (
            r"\bincorporated\b",
            "registration",
        ),
        (
            r"\bseed\s+funding\b",
            "seed fund",
        ),
        (
            r"\bgovt\b",
            "government",
        ),
    )

    for pattern, replacement in replacements:

        text = re.sub(
            pattern,
            replacement,
            text,
        )

    return " ".join(
        text.split()
    )


def _service_title_metrics(
    incoming,
    existing,
):

    left = _canonical_service_title(
        incoming
    )

    right = _canonical_service_title(
        existing
    )

    if (
        not left
        or not right
    ):

        return {
            "score": 0,
            "sequence": 0,
            "containment": 0,
            "jaccard": 0,
            "distinctive_shared": 0,
            "exact": False,
            "anchor": None,
        }

    if left == right:

        return {
            "score": 1,
            "sequence": 1,
            "containment": 1,
            "jaccard": 1,
            "distinctive_shared": 99,
            "exact": True,
            "anchor": "exact",
        }

    left_tokens = set(
        left.split()
    )

    right_tokens = set(
        right.split()
    )

    shared = (
        left_tokens
        & right_tokens
    )

    union = (
        left_tokens
        | right_tokens
    )

    sequence = SequenceMatcher(
        None,
        left,
        right,
    ).ratio()

    containment = (
        len(shared)
        / min(
            len(left_tokens),
            len(right_tokens),
        )
        if (
            left_tokens
            and right_tokens
        )
        else 0
    )

    jaccard = (
        len(shared)
        / len(union)
        if union
        else 0
    )

    generic_tokens = {
        "service",
        "services",
        "scheme",
        "application",
        "company",
        "registration",
        "certificate",
        "certification",
        "government",
        "india",
        "indian",
        "fund",
        "funding",
        "grant",
        "loan",
        "programme",
        "program",
    }

    distinctive_shared = (
        shared
        - generic_tokens
    )

    combined = max(
        sequence,
        (
            0.55
            * containment
            + 0.45
            * jaccard
        ),
    )

    anchor = None


    # --------------------------------------------------------
    # STRONG BUSINESS-IDENTITY ANCHORS
    # --------------------------------------------------------

    phrase_anchors = (
        (
            "limited liability partnership",
            0.97,
        ),
        (
            "80iac",
            0.97,
        ),
        (
            "seed fund",
            0.93,
        ),
        (
            "startup india",
            0.90,
        ),
    )

    for phrase, floor in phrase_anchors:

        if (
            phrase in left
            and phrase in right
        ):

            combined = max(
                combined,
                floor,
            )

            anchor = phrase

            break


    # --------------------------------------------------------
    # STRUCTURED BUSINESS-IDENTITY COMBINATIONS
    #
    # "Private Limited" alone is too broad because many
    # compliance services contain those words.
    # Registration + Private Limited together is meaningful.
    # --------------------------------------------------------

    if (
        "private limited" in left
        and "private limited" in right
        and "registration" in left
        and "registration" in right
    ):

        combined = max(
            combined,
            0.95,
        )

        anchor = (
            "private limited registration"
        )


    # --------------------------------------------------------
    # CONTAINED TITLES
    #
    # Example:
    # Startup India
    # Startup India Recognition
    # --------------------------------------------------------

    smaller = min(
        len(left_tokens),
        len(right_tokens),
    )

    if (
        smaller >= 2
        and containment == 1
        and distinctive_shared
    ):

        combined = max(
            combined,
            0.88,
        )


    return {
        "score":
            min(
                round(
                    combined,
                    4,
                ),
                1,
            ),

        "sequence":
            round(
                sequence,
                4,
            ),

        "containment":
            round(
                containment,
                4,
            ),

        "jaccard":
            round(
                jaccard,
                4,
            ),

        "distinctive_shared":
            len(
                distinctive_shared
            ),

        "exact":
            False,

        "anchor":
            anchor,
    }


def match_service(
    title,
    services,
):

    normalized = _canonical_service_title(
        title
    )

    if not normalized:

        return {
            "action":
                "INVALID",

            "matched_service_id":
                None,

            "score":
                0,

            "method":
                "EMPTY_TITLE",
        }


    candidates = []


    for (
        service_id,
        service_title,
    ) in services:

        metrics = (
            _service_title_metrics(
                title,
                service_title,
            )
        )

        candidates.append(
            (
                metrics["score"],
                metrics[
                    "distinctive_shared"
                ],
                metrics[
                    "containment"
                ],
                service_id,
                service_title,
                metrics,
            )
        )


    candidates.sort(
        key=lambda item: (
            item[0],
            item[1],
            item[2],
        ),
        reverse=True,
    )


    if not candidates:

        return {
            "action":
                "CREATE",

            "matched_service_id":
                None,

            "score":
                0,

            "method":
                "NO_EXISTING_SERVICES",
        }


    (
        best_score,
        distinctive_shared,
        containment,
        service_id,
        service_title,
        metrics,
    ) = candidates[0]


    # --------------------------------------------------------
    # EXACT / SEMANTICALLY CLEAR EXISTING SERVICE
    # --------------------------------------------------------

    if metrics["exact"]:

        return {
            "action":
                "UPDATE",

            "matched_service_id":
                service_id,

            "score":
                1,

            "method":
                "EXACT_NORMALIZED_TITLE",

            "matched_title":
                service_title,
        }


    if (
        best_score >= 0.90
        or (
            best_score >= 0.82
            and distinctive_shared >= 2
        )
    ):

        return {
            "action":
                "UPDATE",

            "matched_service_id":
                service_id,

            "score":
                best_score,

            "method":
                (
                    "BUSINESS_IDENTITY_MATCH"
                    if metrics["anchor"]
                    else "STRONG_TITLE_MATCH"
                ),

            "matched_title":
                service_title,
        }


    # --------------------------------------------------------
    # POSSIBLE EXISTING SERVICE
    #
    # Never silently create when the title is reasonably close.
    # --------------------------------------------------------

    if best_score >= 0.72:

        return {
            "action":
                "MERGE_REVIEW",

            "matched_service_id":
                service_id,

            "score":
                best_score,

            "method":
                "POSSIBLE_EXISTING_SERVICE",

            "matched_title":
                service_title,
        }


    return {
        "action":
            "CREATE",

        "matched_service_id":
            None,

        "score":
            best_score,

        "method":
            "NO_SAFE_EXISTING_MATCH",
    }



def _header_value(
    rows,
    row_number,
    column_index,
):

    for (
        current_row,
        values,
        links,
    ) in rows:

        if current_row != row_number:
            continue

        if (
            column_index
            < len(values)
        ):

            return str(
                values[
                    column_index
                ]
                or ""
            ).strip()

        return ""

    return ""


def suspicious_detected_header(
    rows,
    header,
    services,
):

    """
    Detect a false header selected from actual data.

    Example pattern:
    - first data starts on row 1
    - detector chooses row 9
    - row 9 contains an actual Service title
    - only one/two fields were supposedly recognised

    This is generic and does not depend on a sheet name.
    """

    if (
        not header
        or not rows
    ):

        return False


    mapping = (
        header.get(
            "mapping",
            {}
        )
        or {}
    )

    scheme_column = (
        mapping.get(
            "scheme_name"
        )
    )

    if scheme_column is None:

        return False


    header_row_number = (
        header.get(
            "row"
        )
    )

    if header_row_number is None:

        return False


    value = _header_value(
        rows,
        header_row_number,
        scheme_column,
    )


    # Genuine headers commonly use these words.
    normalized = norm(
        value
    )

    genuine_header_labels = {
        "scheme",
        "scheme name",
        "service",
        "services",
        "service name",
        "program",
        "programme",
        "program name",
        "programme name",
        "title",
    }


    if normalized in genuine_header_labels:

        return False


    match = match_service(
        value,
        services,
    )


    # A supposed header cell that exactly/strongly identifies
    # an existing Service is almost certainly a data row.
    if (
        match.get(
            "matched_service_id"
        )
        and match.get(
            "action"
        )
        == "UPDATE"
        and float(
            match.get(
                "score",
                0,
            )
            or 0
        )
        >= 0.90
    ):

        return True


    first_source_row = min(
        row_number
        for (
            row_number,
            values,
            links,
        )
        in rows
    )


    # Conservative fallback:
    # a late "header" with only one/two recognised fields is
    # suspicious when meaningful data started much earlier.
    if (
        header_row_number
        > first_source_row
        and len(
            mapping
        )
        <= 2
    ):

        return True


    return False


def _financial_content_score(
    value,
):

    text = str(
        value or ""
    ).lower()

    if not text:

        return 0


    score = 0


    markers = (
        "₹",
        "inr",
        "rs.",
        "rs ",
        "lakh",
        "crore",
        "grant",
        "grant-in-aid",
        "subsidy",
        "equity",
        "revenue share",
        "financial support",
        "funding support",
        "support up to",
    )


    for marker in markers:

        if marker in text:

            score += 2


    if "%" in text:

        score += 2


    if re.search(
        r"\b\d+\s*(?:lakh|crore)\b",
        text,
    ):

        score += 3


    return score


def _eligibility_content_score(
    value,
):

    text = norm(
        value
    )

    if not text:

        return 0


    score = 0


    markers = (
        "eligible",
        "eligibility",
        "applicant",
        "registered",
        "incorporated",
        "pvt",
        "private limited",
        "llp",
        "partnership",
        "proprietorship",
        "individual",
        "citizen",
        "must have",
        "should not",
        "should be",
        "company should",
        "startup should",
        "start up should",
        "years in existence",
        "program is open",
        "programme is open",
        "for profit",
        "non profit",
        "nonprofit",
        "fcra",
        "license",
        "licence",
        "compulsory",
    )


    for marker in markers:

        if marker in text:

            score += 2


    return score


def _sector_content_score(
    value,
):

    text = norm(
        value
    )

    raw = str(
        value or ""
    )


    if not text:

        return 0


    score = 0


    markers = (
        "agriculture",
        "agritech",
        "agri tech",
        "healthcare",
        "health tech",
        "medtech",
        "manufacturing",
        "energy",
        "iot",
        "cleantech",
        "clean tech",
        "fintech",
        "edtech",
        "ed tech",
        "saas",
        "education",
        "environment",
        "livelihood",
        "dairy",
        "food tech",
        "food technology",
        "electronics",
        "electrical",
        "mechanical",
        "toys",
        "technology",
        "social impact",
        "transportation",
        "infrastructure",
        "ev",
        "deep tech",
        "green energy",
    )


    for marker in markers:

        if marker in text:

            score += 2


    # Sector lists frequently contain several short lines.
    if raw.count(
        "\n"
    ) >= 2:

        score += 2


    return score


def _append_candidate_field(
    fields,
    field_name,
    value,
):

    value = safe(
        value
    )

    if value in {
        None,
        "",
    }:

        return


    if field_name not in fields:

        fields[
            field_name
        ] = value

        return


    existing = str(
        fields[
            field_name
        ]
    ).strip()

    incoming = str(
        value
    ).strip()


    if (
        not incoming
        or incoming in existing
    ):

        return


    fields[
        field_name
    ] = (
        existing
        + "\n"
        + incoming
    )


def headerless_structured_candidates(
    rows,
    available,
    services,
):

    """
    Generic parser for structured tables where no reliable
    header exists.

    Strategy:
    - identify Service title per row
    - classify other cells using content evidence
    - never invent Funding Organisation without an explicit
      header
    - preserve uncertain business text as Additional Info
    """

    result = []


    for (
        row_no,
        values,
        links,
    ) in rows:

        populated = [
            (
                index,
                value,
            )
            for (
                index,
                value,
            )
            in enumerate(
                values
            )
            if str(
                value or ""
            ).strip()
        ]


        if not populated:

            continue


        # ----------------------------------------------------
        # FIND SERVICE TITLE
        # ----------------------------------------------------

        title_index = None

        title = ""

        best_match = None


        for index, value in populated:

            candidate_title = str(
                value
            ).strip()


            # Descriptive paragraphs are not good titles.
            if (
                len(
                    candidate_title
                )
                > 180
                or "\n"
                in candidate_title
            ):

                continue


            match = match_service(
                candidate_title,
                services,
            )


            if (
                match.get(
                    "matched_service_id"
                )
                and match.get(
                    "action"
                )
                == "UPDATE"
            ):

                score = float(
                    match.get(
                        "score",
                        0,
                    )
                    or 0
                )


                if (
                    best_match is None
                    or score
                    > best_match[
                        0
                    ]
                ):

                    best_match = (
                        score,
                        index,
                        candidate_title,
                        match,
                    )


        if best_match:

            (
                _,
                title_index,
                title,
                title_match,
            ) = best_match


        else:

            # Future new Services may not exist in DB yet.
            # Prefer the first concise non-paragraph cell.
            for index, value in populated:

                candidate_title = str(
                    value
                ).strip()


                if (
                    len(
                        candidate_title
                    )
                    <= 140
                    and "\n"
                    not in candidate_title
                    and len(
                        candidate_title.split()
                    )
                    <= 18
                ):

                    title_index = index
                    title = candidate_title
                    break


            if title_index is None:

                continue


            title_match = match_service(
                title,
                services,
            )


        fields = {
            "scheme_name":
                title,
        }

        source_links = []


        # ----------------------------------------------------
        # CLASSIFY REMAINING BUSINESS CELLS
        # ----------------------------------------------------

        for (
            index,
            value,
        ) in populated:

            if index == title_index:

                continue


            financial_score = (
                _financial_content_score(
                    value
                )
            )

            eligibility_score = (
                _eligibility_content_score(
                    value
                )
            )

            sector_score = (
                _sector_content_score(
                    value
                )
            )


            scores = [
                (
                    financial_score,
                    "benefits",
                ),
                (
                    eligibility_score,
                    "eligibility",
                ),
                (
                    sector_score,
                    "focus_sectors",
                ),
            ]


            scores.sort(
                reverse=True
            )


            best_score, best_field = (
                scores[0]
            )


            # Require actual semantic evidence.
            if best_score >= 2:

                _append_candidate_field(
                    fields,
                    best_field,
                    value,
                )


            else:

                _append_candidate_field(
                    fields,
                    "additional_info",
                    value,
                )


            if (
                index
                < len(
                    links
                )
                and links[
                    index
                ]
            ):

                source_links.append(
                    {
                        "field":
                            best_field
                            if best_score
                            >= 2
                            else "additional_info",

                        "url":
                            links[
                                index
                            ],
                    }
                )


        combined = (
            title
            + " "
            + " ".join(
                str(
                    value
                )
                for value
                in fields.values()
            )
        )


        result.append(
            {
                "source_row":
                    row_no,

                "title":
                    title,

                "fields":
                    fields,

                "links":
                    source_links,

                "proposal":
                    propose(
                        combined,
                        available,
                    ),

                "match":
                    title_match,

                "extraction_mode":
                    "HEADERLESS_STRUCTURED",
            }
        )


    return result


def structured_candidates(
    rows,
    header,
    available,
    services,
):

    result = []

    mapping = header[
        "mapping"
    ]

    for (
        row_no,
        values,
        links,
    ) in rows:

        if (
            row_no
            <= header["row"]
        ):
            continue

        fields = {}
        source_links = []

        for (
            field,
            index,
        ) in mapping.items():

            if (
                index < len(values)
                and str(
                    values[index]
                ).strip()
            ):
                fields[field] = safe(
                    values[index]
                )

            if (
                index < len(links)
                and links[index]
            ):
                source_links.append(
                    {
                        "field":
                            field,

                        "url":
                            links[index],
                    }
                )

        title = str(
            fields.get(
                "scheme_name",
                "",
            )
        ).strip()

        if not title:
            continue

        combined = (
            title
            + " "
            + " ".join(
                str(value)
                for value
                in fields.values()
            )
        )

        result.append(
            {
                "source_row":
                    row_no,

                "title":
                    title,

                "fields":
                    fields,

                "links":
                    source_links,

                "proposal":
                    propose(
                        combined,
                        available,
                    ),

                "match":
                    match_service(
                        title,
                        services,
                    ),
            }
        )

    return result



def narrative_candidate(
    sheet_name,
    rows,
    available,
    services,
):

    sections = []

    seen = set()

    current = "OTHER"

    title_hints = []


    def clean_title(value):

        return re.sub(
            r"[*_]+",
            "",
            str(
                value or ""
            ),
        ).strip(
            " :-•▪️◼️"
        )


    def looks_like_service_title(
        value,
    ):

        value = clean_title(
            value
        )

        if not value:

            return False

        if (
            len(value) < 4
            or len(value) > 150
        ):

            return False

        if (
            value.count(",") > 2
            or value.count(";") > 1
        ):

            return False

        words = value.split()

        if (
            len(words) < 2
            or len(words) > 18
        ):

            return False

        normalized = norm(
            value
        )

        if normalized in {
            "dear all",
            "please note",
            "benefits",
            "eligibility",
            "documents required",
            "scope of work",
            "commercials",
        }:

            return False

        markers = (
            "registration",
            "scheme",
            "startup",
            "start up",
            "seed fund",
            "tax exemption",
            "80 iac",
            "80iac",
            "certificate",
            "certification",
            "incorporation",
            "licence",
            "license",
            "loan",
            "grant",
            "subsidy",
            "funding",
        )

        return any(
            marker in normalized
            for marker in markers
        )


    # --------------------------------------------------------
    # SHEET NAME IS A HINT, NOT AUTOMATIC TRUTH
    # --------------------------------------------------------

    sheet_hint = clean_title(
        str(
            sheet_name
        ).replace(
            "_",
            " ",
        )
    )

    if sheet_hint:

        title_hints.append(
            sheet_hint
        )


    # --------------------------------------------------------
    # PRESERVE EVERY NARRATIVE SECTION
    # --------------------------------------------------------

    for (
        row_no,
        values,
        links,
    ) in rows:

        for (
            column,
            value,
        ) in enumerate(
            values,
            1,
        ):

            content = str(
                value or ""
            ).strip()

            key = (
                row_no,
                column,
                norm(
                    content
                ),
            )

            if (
                not content
                or key in seen
            ):

                continue

            seen.add(
                key
            )

            heading = is_heading(
                content
            )

            detected = section_type(
                content
            )

            if (
                heading
                and detected
            ):

                current = detected

            effective_type = (
                detected
                if (
                    heading
                    and detected
                )
                else current
            )

            url = (
                links[
                    column - 1
                ]
                if (
                    column - 1
                    < len(links)
                )
                else None
            )

            sections.append(
                {
                    "source_row":
                        row_no,

                    "source_column":
                        column,

                    "content":
                        content,

                    "section_type":
                        effective_type,

                    "is_heading":
                        bool(
                            heading
                        ),

                    "url":
                        url,
                }
            )

            if looks_like_service_title(
                content
            ):

                cleaned = clean_title(
                    content
                )

                if (
                    cleaned
                    not in title_hints
                ):

                    title_hints.append(
                        cleaned
                    )


    # --------------------------------------------------------
    # FIND THE BEST EXISTING SERVICE USING ALL TITLE HINTS
    # --------------------------------------------------------

    service_titles = {
        service_id:
            service_title
        for (
            service_id,
            service_title,
        ) in services
    }

    selected_hint = None
    selected_match = None


    # --------------------------------------------------------
    # A. TRUST A STRONG SHEET-SUBJECT MATCH FIRST
    #
    # Example:
    # START_UP INDIA -> Startup India Recognition
    # SEED FUND      -> Startup India Seed Fund Scheme
    # LLP            -> LLP Incorporation
    #
    # Once this is strong, later references inside the sheet
    # cannot replace the sheet's main subject.
    # --------------------------------------------------------

    sheet_match = None

    sheet_hint_tokens = (
        re.findall(
            r"[A-Za-z0-9]+",
            sheet_hint or "",
        )
    )

    sheet_hint_is_specific = (
        len(
            sheet_hint_tokens
        )
        >= 3
    )


    if sheet_hint:

        sheet_match = match_service(
            sheet_hint,
            services,
        )


        # ----------------------------------------------------
        # A short sheet label such as:
        #
        # LLP
        # PVT
        # SEED FUND
        # TAX CERT
        #
        # describes a topic/family and must NOT automatically
        # override the actual Service title found inside the
        # sheet.
        #
        # However an exact Service title is always safe.
        # Longer specific sheet subjects such as
        # "START UP INDIA" may also be trusted when strongly
        # matched.
        # ----------------------------------------------------

        sheet_match_is_exact = (
            sheet_match.get(
                "method"
            )
            == "EXACT_NORMALIZED_TITLE"
        )


        if (
            sheet_match.get("action")
            == "UPDATE"
            and (
                sheet_match_is_exact
                or sheet_hint_is_specific
            )
        ):

            selected_hint = (
                sheet_hint
            )

            selected_match = (
                sheet_match
            )


    # --------------------------------------------------------
    # B. OTHERWISE USE THE FIRST STRONG SERVICE TITLE
    #
    # Source order matters. The first genuine title/heading is
    # normally the subject; later headings may describe related
    # services and must not hijack the match.
    # --------------------------------------------------------

    content_hints = (
        title_hints[1:]
        if (
            title_hints
            and title_hints[0]
            == sheet_hint
        )
        else title_hints
    )


    if selected_match is None:

        first_possible = None

        first_create = None

        for hint in content_hints:

            match = match_service(
                hint,
                services,
            )

            action = match.get(
                "action"
            )

            if action == "UPDATE":

                selected_hint = hint
                selected_match = match
                break

            if (
                action == "MERGE_REVIEW"
                and first_possible is None
            ):

                first_possible = (
                    hint,
                    match,
                )

            if (
                action == "CREATE"
                and first_create is None
            ):

                first_create = (
                    hint,
                    match,
                )


        # Prefer sheet-name possible match over later prose.
        if selected_match is None:

            sheet_possible = None

            if sheet_hint:

                candidate = match_service(
                    sheet_hint,
                    services,
                )

                if (
                    candidate.get("action")
                    == "MERGE_REVIEW"
                ):

                    sheet_possible = (
                        sheet_hint,
                        candidate,
                    )


            chosen = (
                sheet_possible
                or first_possible
                or first_create
            )

            if chosen:

                (
                    selected_hint,
                    selected_match,
                ) = chosen


    if selected_match is None:

        selected_hint = (
            sheet_hint
            or str(
                sheet_name
            )
        )

        selected_match = {
            "action":
                "INVALID",

            "matched_service_id":
                None,

            "score":
                0,

            "method":
                "NO_SERVICE_TITLE_FOUND",
        }


    matched_id = (
        selected_match.get(
            "matched_service_id"
        )
    )


    # --------------------------------------------------------
    # IF KNOWLEDGE CLEARLY BELONGS TO AN EXISTING SERVICE,
    # DISPLAY THE REAL TOOLKIT SERVICE TITLE.
    # --------------------------------------------------------

    if (
        matched_id
        and selected_match.get(
            "action"
        )
        in {
            "UPDATE",
            "MERGE_REVIEW",
        }
    ):

        title = (
            service_titles.get(
                matched_id
            )
            or selected_hint
        )

    else:

        title = selected_hint


    # --------------------------------------------------------
    # EXTRA PROTECTION AGAINST DESCRIPTIVE SENTENCES
    #
    # Narrative prose should never become CREATE merely because
    # it happened to contain words such as "funding" or "scheme".
    # --------------------------------------------------------

    if (
        selected_match.get(
            "action"
        )
        == "CREATE"
        and not looks_like_service_title(
            selected_hint
        )
    ):

        selected_match = {
            **selected_match,

            "action":
                "INVALID",

            "method":
                "NARRATIVE_TITLE_NOT_SERVICE_LIKE",
        }


    combined_text = " ".join(
        section[
            "content"
        ]
        for section in sections
    )


    return {
        "source_row":
            (
                sections[0][
                    "source_row"
                ]
                if sections
                else 1
            ),

        "title":
            title,

        "source_title_hint":
            selected_hint,

        "sections":
            sections,

        "proposal":
            propose(
                (
                    title
                    + " "
                    + combined_text
                ),
                available,
            ),

        "match":
            selected_match,
    }


def xlsx_sheets(
    file_obj,
):

    file_obj.seek(0)

    workbook = load_workbook(
        file_obj,
        read_only=False,
        data_only=False,
    )

    result = []

    try:

        for worksheet in (
            workbook.worksheets
        ):

            rows = []

            max_column = min(
                worksheet.max_column
                or 1,
                200,
            )

            for row in worksheet.iter_rows(
                min_row=1,
                max_row=(
                    worksheet.max_row
                    or 1
                ),
                max_col=max_column,
            ):

                values = [
                    safe(cell.value)
                    for cell in row
                ]

                links = [
                    (
                        cell.hyperlink.target
                        if cell.hyperlink
                        else None
                    )
                    for cell in row
                ]

                if any(
                    str(value).strip()
                    or link
                    for (
                        value,
                        link,
                    )
                    in zip(
                        values,
                        links,
                    )
                ):

                    while (
                        values
                        and not str(
                            values[-1]
                        ).strip()
                        and not links[-1]
                    ):
                        values.pop()
                        links.pop()

                    rows.append(
                        (
                            row[0].row,
                            values,
                            links,
                        )
                    )

            result.append(
                (
                    worksheet.title,
                    rows,
                )
            )

    finally:

        workbook.close()

        file_obj.seek(0)

    return result


def csv_sheets(
    file_obj,
):

    file_obj.seek(0)

    raw = file_obj.read()

    file_obj.seek(0)

    if isinstance(
        raw,
        bytes,
    ):

        try:
            text = raw.decode(
                "utf-8-sig"
            )

        except UnicodeDecodeError:

            text = raw.decode(
                "cp1252"
            )

    else:
        text = raw

    try:

        dialect = csv.Sniffer().sniff(
            text[:8192],
            delimiters=",;\t|",
        )

    except csv.Error:

        dialect = csv.excel

    rows = []

    for (
        row_no,
        values,
    ) in enumerate(
        csv.reader(
            io.StringIO(text),
            dialect,
        ),
        1,
    ):

        values = [
            safe(value)
            for value
            in values
        ]

        if any(
            str(value).strip()
            for value in values
        ):

            rows.append(
                (
                    row_no,
                    values,
                    [None] * len(
                        values
                    ),
                )
            )

    return [
        (
            "CSV",
            rows,
        )
    ]


def analyse_file(
    file_obj,
    filename,
):

    extension = Path(
        filename
    ).suffix.lower()

    if extension not in {
        ".xlsx",
        ".csv",
    }:

        raise ValueError(
            "Only .xlsx and .csv files are supported."
        )

    available = set(
        Category.objects
        .values_list(
            "slug",
            flat=True,
        )
    )

    services = list(
        Service.objects
        .values_list(
            "id",
            "title",
        )
    )

    if extension == ".xlsx":

        source_sheets = (
            xlsx_sheets(
                file_obj
            )
        )

    else:

        source_sheets = (
            csv_sheets(
                file_obj
            )
        )

    sheets = []

    action_counts = Counter()
    kind_counts = Counter()

    total_candidates = 0

    for (
        sheet_name,
        rows,
    ) in source_sheets:

        header = detect_header(
            rows
        )

        kind = sheet_kind(
            rows,
            header,
        )

        # ----------------------------------------------------
        # HEADERLESS RECOVERY SAFETY
        #
        # Only a sheet that was ALREADY recognised as a
        # structured table may enter headerless recovery.
        #
        # Never reinterpret:
        # - narrative knowledge
        # - comparison matrices
        # - communications
        # - reference tables
        #
        # as Service-row tables.
        # ----------------------------------------------------

        headerless_structured = (
            kind
            == "STRUCTURED_TABLE"
            and suspicious_detected_header(
                rows,
                header,
                services,
            )
        )

        kind_counts[
            kind
        ] += 1

        candidates = []

        if headerless_structured:

            candidates = (
                headerless_structured_candidates(
                    rows,
                    available,
                    services,
                )
            )

        elif (
            kind in {
                "STRUCTURED_TABLE",
                "COMMERCIAL_SERVICE_TABLE",
            }
            and header
        ):

            candidates = (
                structured_candidates(
                    rows,
                    header,
                    available,
                    services,
                )
            )

        elif (
            kind
            == "NARRATIVE_KNOWLEDGE"
        ):

            candidate = (
                narrative_candidate(
                    sheet_name,
                    rows,
                    available,
                    services,
                )
            )

            if candidate[
                "sections"
            ]:

                candidates = [
                    candidate
                ]

        for candidate in (
            candidates
        ):

            action_counts[
                candidate[
                    "match"
                ][
                    "action"
                ]
            ] += 1

        total_candidates += len(
            candidates
        )

        sheets.append(
            {
                "name":
                    sheet_name,

                "kind":
                    kind,

                "nonempty_rows":
                    len(rows),

                "header_row":
                    (
                        header["row"]
                        if header
                        else None
                    ),

                "field_mapping":
                    (
                        header[
                            "mapping"
                        ]
                        if header
                        else {}
                    ),

                "candidate_count":
                    len(
                        candidates
                    ),

                "candidates":
                    candidates,
            }
        )

    return {
        "engine_version":
            "1.1",

        "business_knowledge_visibility":
            "BDE",

        "commercial_business_data_visible_to_bde":
            True,

        "technical_import_internals_visible_to_bde":
            False,

        "source_type":
            (
                "XLSX"
                if extension
                == ".xlsx"
                else "CSV"
            ),

        "source_name":
            Path(
                filename
            ).name,

        "file_sha256":
            sha256(
                file_obj
            ),

        "sheet_count":
            len(
                sheets
            ),

        "sheet_type_counts":
            dict(
                kind_counts
            ),

        "candidate_count":
            total_candidates,

        "candidate_action_counts":
            dict(
                action_counts
            ),

        "sheets":
            sheets,

        "database_writes":
            0,
    }
