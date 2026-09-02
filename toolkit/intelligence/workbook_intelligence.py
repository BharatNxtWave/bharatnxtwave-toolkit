from __future__ import annotations

import hashlib
import json
import re

from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.apps import apps
from django.db.models import Q
from openpyxl import load_workbook


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

WORKBOOK_PATH = (
    PROJECT_ROOT
    / "confidential_source"
    / "BHARATNXT  WAVE_POWER _ TOOLKIT.xlsx"
)

AUDIT_DIR = (
    PROJECT_ROOT
    / "confidential_source"
    / "audit"
)

AUDIT_DIR.mkdir(
    parents=True,
    exist_ok=True,
)


# ============================================================
# SECURITY / VISIBILITY
# ============================================================

BLOCKED_KEYWORDS = {
    "vendor",
    "bdm",
    "deduction",
    "internal",
    "escalation",
    "commercial",
    "commission",
    "margin",
    "employee",
    "password",
}


SECTION_PATTERNS = {
    "ELIGIBILITY": (
        "eligibility",
        "eligible",
        "eligibility criteria",
        "who can apply",
        "applicant eligibility",
        "criteria",
    ),

    "DOCUMENTS": (
        "document",
        "documents",
        "required documents",
        "documentation",
        "checklist",
    ),

    "PROCESS": (
        "process",
        "procedure",
        "steps",
        "how to apply",
        "application process",
        "application procedure",
    ),

    "BENEFITS": (
        "benefit",
        "benefits",
        "advantage",
        "assistance",
        "support offered",
    ),

    "FUNDING": (
        "funding",
        "finance",
        "financial assistance",
        "loan amount",
        "grant amount",
        "subsidy",
        "fund",
    ),

    "TIMELINE": (
        "timeline",
        "deadline",
        "last date",
        "closing date",
        "application date",
        "duration",
        "tenure",
    ),

    "SCOPE": (
        "scope",
        "applicable",
        "applicable for",
        "sector",
        "industry",
        "industries",
    ),

    "OVERVIEW": (
        "overview",
        "description",
        "about",
        "scheme details",
        "details",
    ),

    "NOTES": (
        "note",
        "notes",
        "important",
        "remark",
        "remarks",
    ),
}


# ============================================================
# GENERIC HELPERS
# ============================================================

def json_safe(value):

    if value is None:
        return None

    if isinstance(
        value,
        (
            str,
            int,
            float,
            bool,
        ),
    ):
        return value

    if isinstance(
        value,
        (
            datetime,
            date,
            Decimal,
        ),
    ):
        return str(value)

    if isinstance(value, dict):
        return {
            str(k): json_safe(v)
            for k, v
            in value.items()
        }

    if isinstance(
        value,
        (
            list,
            tuple,
            set,
        ),
    ):
        return [
            json_safe(v)
            for v in value
        ]

    return str(value)


def text(value):

    if value is None:
        return ""

    value = str(value).strip()

    if value.lower() in {
        "none",
        "null",
        "nan",
    }:
        return ""

    return value


def normalise_key(value):

    value = text(value).lower()

    value = re.sub(
        r"[_\-/]+",
        " ",
        value,
    )

    value = re.sub(
        r"\s+",
        " ",
        value,
    )

    return value.strip()


def blocked_key(key):

    key = normalise_key(key)

    return any(
        token in key
        for token
        in BLOCKED_KEYWORDS
    )


def unique_strings(values):

    seen = set()
    result = []

    for value in values:

        value = text(value)

        if not value:
            continue

        fingerprint = (
            re.sub(
                r"\s+",
                " ",
                value,
            )
            .strip()
            .lower()
        )

        if fingerprint in seen:
            continue

        seen.add(fingerprint)
        result.append(value)

    return result


def sha256_file(path):

    digest = hashlib.sha256()

    with open(path, "rb") as handle:

        while True:

            block = handle.read(
                1024 * 1024
            )

            if not block:
                break

            digest.update(block)

    return digest.hexdigest()


# ============================================================
# PAYLOAD EXTRACTION
# ============================================================

def payload_value(payload):

    if not isinstance(
        payload,
        dict,
    ):
        return text(payload)

    return text(
        payload.get("value")
    )


def extract_payload_urls(payload):

    if not isinstance(
        payload,
        dict,
    ):
        return []

    values = []

    for key in (
        "hyperlink_target",
        "hyperlink_location",
    ):

        candidate = text(
            payload.get(key)
        )

        if candidate:
            values.append(candidate)

    value = payload_value(
        payload
    )

    if value:

        for match in re.findall(
            r'https?://[^\s"\'<>]+',
            value,
            flags=re.I,
        ):
            values.append(match)

    return unique_strings(values)


def flatten_raw_data(
    obj,
    prefix="",
):

    result = []

    if isinstance(obj, dict):

        # Special handling for staged cell payload.
        if "value" in obj:

            value = payload_value(
                obj
            )

            if value:

                result.append(
                    {
                        "key": prefix,
                        "value": value,
                        "urls": (
                            extract_payload_urls(
                                obj
                            )
                        ),
                    }
                )

            return result

        for key, value in obj.items():

            if key == "_meta":
                continue

            next_prefix = (
                f"{prefix}.{key}"
                if prefix
                else str(key)
            )

            result.extend(
                flatten_raw_data(
                    value,
                    next_prefix,
                )
            )

    elif isinstance(
        obj,
        (
            list,
            tuple,
        ),
    ):

        for index, value in enumerate(obj):

            result.extend(
                flatten_raw_data(
                    value,
                    f"{prefix}[{index}]",
                )
            )

    else:

        value = text(obj)

        if value:

            result.append(
                {
                    "key": prefix,
                    "value": value,
                    "urls": [],
                }
            )

    return result


# ============================================================
# SECTION CLASSIFICATION
# ============================================================

def detect_section_type(key):

    key = normalise_key(key)

    if not key:
        return None

    if blocked_key(key):
        return None

    # Most specific checks first.
    for section_type in (
        "DOCUMENTS",
        "ELIGIBILITY",
        "PROCESS",
        "TIMELINE",
        "FUNDING",
        "BENEFITS",
        "SCOPE",
        "OVERVIEW",
        "NOTES",
    ):

        for phrase in (
            SECTION_PATTERNS[
                section_type
            ]
        ):

            if phrase in key:
                return section_type

    return None


# ============================================================
# LIST SPLITTING
# Conservative: only clear list structures.
# ============================================================

LIST_PREFIX = re.compile(
    r"""
    ^\s*
    (?:
        [•●▪◦\-–—*]
        |
        \d+[\.\)]
        |
        [a-zA-Z][\.\)]
    )
    \s+
    """,
    re.X,
)


def split_explicit_items(value):

    value = text(value)

    if not value:
        return []

    lines = [
        line.strip()
        for line
        in value.splitlines()
        if line.strip()
    ]

    explicit = []

    for line in lines:

        if LIST_PREFIX.search(line):

            cleaned = (
                LIST_PREFIX.sub(
                    "",
                    line,
                )
                .strip()
            )

            if cleaned:
                explicit.append(cleaned)

    if explicit:
        return unique_strings(
            explicit
        )

    # If source uses pipe separators,
    # treat as structured only when several parts exist.
    if "|" in value:

        parts = [
            item.strip()
            for item
            in value.split("|")
            if item.strip()
        ]

        if len(parts) >= 2:
            return unique_strings(parts)

    # Preserve the original source as ONE rule/item.
    return [value]


# ============================================================
# WORKBOOK CENSUS
# ============================================================

def workbook_census():

    if not WORKBOOK_PATH.exists():

        raise RuntimeError(
            f"Workbook not found: "
            f"{WORKBOOK_PATH}"
        )

    workbook = load_workbook(
        WORKBOOK_PATH,
        read_only=False,
        data_only=False,
        keep_links=True,
    )

    sheet_data = []

    total_nonempty = 0
    total_hyperlinks = 0
    total_formulas = 0
    total_merged = 0

    for ws in workbook.worksheets:

        nonempty = 0
        hyperlinks = 0
        formulas = 0

        for row in ws.iter_rows():

            for cell in row:

                value = cell.value

                if (
                    value is not None
                    and str(value).strip()
                ):
                    nonempty += 1

                if cell.hyperlink:
                    hyperlinks += 1

                if (
                    isinstance(value, str)
                    and value.startswith("=")
                ):
                    formulas += 1

        merged = len(
            ws.merged_cells.ranges
        )

        total_nonempty += nonempty
        total_hyperlinks += hyperlinks
        total_formulas += formulas
        total_merged += merged

        sheet_data.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_cells": nonempty,
                "hyperlinks": hyperlinks,
                "formulas": formulas,
                "merged_ranges": merged,
            }
        )

    workbook.close()

    return {
        "path": str(
            WORKBOOK_PATH
        ),
        "sha256": sha256_file(
            WORKBOOK_PATH
        ),
        "sheet_count": len(
            sheet_data
        ),
        "nonempty_cells": (
            total_nonempty
        ),
        "hyperlinks": (
            total_hyperlinks
        ),
        "formulas": (
            total_formulas
        ),
        "merged_ranges": (
            total_merged
        ),
        "sheets": sheet_data,
    }


# ============================================================
# KNOWLEDGE MODEL DISCOVERY
# ============================================================

def discover_knowledge_models():

    discovered = []

    for model in apps.get_app_config(
        "toolkit"
    ).get_models():

        name = model.__name__

        if "knowledge" not in (
            name.lower()
        ):
            continue

        fields = []

        for field in (
            model._meta.get_fields()
        ):

            fields.append(
                {
                    "name": field.name,
                    "type": (
                        type(field).__name__
                    ),
                }
            )

        try:
            count = (
                model.objects.count()
            )
        except Exception:
            count = None

        discovered.append(
            {
                "model": name,
                "count": count,
                "fields": fields,
            }
        )

    return discovered


# ============================================================
# SERVICE SOURCE BUILDING
# ============================================================

def direct_service_sections(service):

    mapping = (
        (
            "OVERVIEW",
            "Overview",
            service.overview,
        ),
        (
            "BENEFITS",
            "Benefits",
            service.benefits,
        ),
        (
            "ELIGIBILITY",
            "Eligibility",
            service.eligibility_summary,
        ),
        (
            "NOTES",
            "Restrictions",
            service.restrictions,
        ),
        (
            "NOTES",
            "Important Notes",
            service.important_notes,
        ),
        (
            "FUNDING",
            "Subsidy Details",
            service.subsidy_details,
        ),
    )

    result = []

    for (
        section_type,
        title,
        value,
    ) in mapping:

        value = text(value)

        if not value:
            continue

        result.append(
            {
                "section_type":
                    section_type,

                "title":
                    title,

                "content":
                    value,

                "source":
                    "SERVICE_FIELD",
            }
        )

    return result


def service_structured_snapshot(
    service,
):

    return {
        "business_types":
            json_safe(
                service.business_types
            ),

        "business_stages":
            json_safe(
                service.business_stages
            ),

        "industries":
            json_safe(
                service.industries
            ),

        "applicable_states":
            json_safe(
                service.applicable_states
            ),

        "founder_categories":
            json_safe(
                service.founder_categories
            ),

        "applicable_for_raw":
            service.applicable_for_raw,

        "min_business_age_months":
            service.min_business_age_months,

        "max_business_age_months":
            service.max_business_age_months,

        "min_turnover":
            json_safe(
                service.min_turnover
            ),

        "max_turnover":
            json_safe(
                service.max_turnover
            ),

        "funding_min":
            json_safe(
                service.funding_min
            ),

        "funding_max":
            json_safe(
                service.funding_max
            ),

        "funding_type":
            service.funding_type,

        "funding_organisation":
            service.funding_organisation,

        "interest_rate_min":
            json_safe(
                service.interest_rate_min
            ),

        "interest_rate_max":
            json_safe(
                service.interest_rate_max
            ),

        "collateral_required":
            service.collateral_required,

        "tenure":
            service.tenure,

        "estimated_processing_time":
            service.estimated_processing_time,

        "application_deadline":
            json_safe(
                service.application_deadline
            ),

        "application_deadline_raw":
            service.application_deadline_raw,

        "deadline_status":
            service.deadline_status,
    }


# ============================================================
# RAW IMPORT ROW EXTRACTION
# ============================================================

def import_rows_for_service(
    ImportRow,
    service,
):

    return (
        ImportRow.objects
        .filter(
            Q(
                imported_service=service
            )
            |
            Q(
                matched_service=service
            )
        )
        .order_by(
            "sheet_name",
            "source_row_number",
        )
    )


def extract_raw_sections(rows):

    candidates = []

    evidence = []

    for row in rows:

        raw = (
            row.raw_data
            if isinstance(
                row.raw_data,
                dict,
            )
            else {}
        )

        flattened = (
            flatten_raw_data(
                raw
            )
        )

        row_evidence = {
            "import_row_id":
                row.id,

            "sheet":
                row.sheet_name,

            "row":
                row.source_row_number,

            "values": [],
        }

        for item in flattened:

            key = item["key"]
            value = item["value"]

            if blocked_key(key):
                continue

            row_evidence[
                "values"
            ].append(
                {
                    "key": key,
                    "value": value,
                    "urls": item["urls"],
                }
            )

            section_type = (
                detect_section_type(
                    key
                )
            )

            if not section_type:
                continue

            candidates.append(
                {
                    "section_type":
                        section_type,

                    "title":
                        normalise_key(
                            key
                        ).title(),

                    "content":
                        value,

                    "source":
                        "IMPORT_ROW",

                    "import_row_id":
                        row.id,

                    "sheet":
                        row.sheet_name,

                    "row":
                        (
                            row
                            .source_row_number
                        ),
                }
            )

        evidence.append(
            row_evidence
        )

    return (
        candidates,
        evidence,
    )


# ============================================================
# DEDUPLICATE SECTION CANDIDATES
# ============================================================

def dedupe_sections(sections):

    seen = set()
    output = []

    for item in sections:

        content = text(
            item.get("content")
        )

        if not content:
            continue

        fingerprint = (
            item.get(
                "section_type",
                "",
            ),
            re.sub(
                r"\s+",
                " ",
                content,
            ).lower(),
        )

        if fingerprint in seen:
            continue

        seen.add(fingerprint)
        output.append(item)

    return output


# ============================================================
# STRUCTURED CANDIDATE GENERATION
# ============================================================

def eligibility_candidates(
    sections,
):

    output = []

    source_sections = [
        item
        for item in sections
        if (
            item["section_type"]
            == "ELIGIBILITY"
        )
    ]

    for section in source_sections:

        items = split_explicit_items(
            section["content"]
        )

        for item in items:

            output.append(
                {
                    "title":
                        "Eligibility criterion",

                    "description":
                        item,

                    "source":
                        section.get(
                            "source"
                        ),

                    "import_row_id":
                        section.get(
                            "import_row_id"
                        ),
                }
            )

    return output


def document_candidates(
    sections,
):

    output = []

    source_sections = [
        item
        for item in sections
        if (
            item["section_type"]
            == "DOCUMENTS"
        )
    ]

    for section in source_sections:

        items = split_explicit_items(
            section["content"]
        )

        for item in items:

            output.append(
                {
                    "name": item,
                    "is_mandatory": None,
                    "source":
                        section.get(
                            "source"
                        ),

                    "import_row_id":
                        section.get(
                            "import_row_id"
                        ),
                }
            )

    return output


def process_candidates(
    sections,
):

    output = []

    source_sections = [
        item
        for item in sections
        if (
            item["section_type"]
            == "PROCESS"
        )
    ]

    counter = 1

    for section in source_sections:

        items = split_explicit_items(
            section["content"]
        )

        for item in items:

            output.append(
                {
                    "step_number":
                        counter,

                    "title":
                        item[:180],

                    "description":
                        item,

                    "source":
                        section.get(
                            "source"
                        ),

                    "import_row_id":
                        section.get(
                            "import_row_id"
                        ),
                }
            )

            counter += 1

    return output


# ============================================================
# COMPLETENESS
# ============================================================

def has_value(value):

    if value is None:
        return False

    if isinstance(
        value,
        (
            list,
            dict,
        ),
    ):
        return bool(value)

    return bool(
        text(value)
    )


def completeness_for_service(
    service,
    sections,
    eligibility,
    documents,
    process,
):

    section_types = {
        item["section_type"]
        for item in sections
    }

    checks = {
        "overview":
            bool(
                text(
                    service.overview
                )
            ),

        "benefits":
            bool(
                text(
                    service.benefits
                )
            ),

        "eligibility":
            bool(
                text(
                    service
                    .eligibility_summary
                )
                or eligibility
            ),

        "documents":
            bool(documents),

        "process":
            bool(process),

        "funding":
            any(
                [
                    service.funding_min
                    is not None,

                    service.funding_max
                    is not None,

                    bool(
                        text(
                            service
                            .funding_type
                        )
                    ),

                    bool(
                        text(
                            service
                            .funding_organisation
                        )
                    ),

                    "FUNDING"
                    in section_types,
                ]
            ),

        "deadline":
            any(
                [
                    service
                    .application_deadline
                    is not None,

                    bool(
                        text(
                            service
                            .application_deadline_raw
                        )
                    ),

                    "TIMELINE"
                    in section_types,
                ]
            ),

        "source_link":
            service.sources.exclude(
                source_url=""
            ).exists(),

        "classification":
            service
            .classification_links
            .exists(),
    }

    available = sum(
        1
        for value
        in checks.values()
        if value
    )

    score = round(
        (
            available
            / len(checks)
        )
        * 100,
        1,
    )

    return {
        "score": score,
        "checks": checks,
        "missing": [
            key
            for key, value
            in checks.items()
            if not value
        ],
    }


# ============================================================
# MAIN PLAN
# ============================================================

def build_intelligence_plan():

    Service = apps.get_model(
        "toolkit",
        "Service",
    )

    ImportRow = apps.get_model(
        "toolkit",
        "ImportRow",
    )

    EligibilityRule = (
        apps.get_model(
            "toolkit",
            "EligibilityRule",
        )
    )

    DocumentRequirement = (
        apps.get_model(
            "toolkit",
            "DocumentRequirement",
        )
    )

    ProcessStep = (
        apps.get_model(
            "toolkit",
            "ProcessStep",
        )
    )

    ServiceContentSection = (
        apps.get_model(
            "toolkit",
            "ServiceContentSection",
        )
    )

    census = workbook_census()

    knowledge_models = (
        discover_knowledge_models()
    )

    services = (
        Service.objects
        .filter(
            status="PUBLISHED"
        )
        .select_related(
            "domain",
            "category",
        )
        .prefetch_related(
            "sources",
            "classification_links",
        )
        .order_by(
            "id"
        )
    )

    service_plans = []

    totals = Counter()

    missing_counts = Counter()

    for service in services:

        rows = list(
            import_rows_for_service(
                ImportRow,
                service,
            )
        )

        (
            raw_sections,
            evidence,
        ) = extract_raw_sections(
            rows
        )

        direct_sections = (
            direct_service_sections(
                service
            )
        )

        sections = (
            dedupe_sections(
                direct_sections
                + raw_sections
            )
        )

        eligibility = (
            eligibility_candidates(
                sections
            )
        )

        documents = (
            document_candidates(
                sections
            )
        )

        process = (
            process_candidates(
                sections
            )
        )

        completeness = (
            completeness_for_service(
                service,
                sections,
                eligibility,
                documents,
                process,
            )
        )

        for missing in (
            completeness["missing"]
        ):
            missing_counts[
                missing
            ] += 1

        totals[
            "sections"
        ] += len(sections)

        totals[
            "eligibility_candidates"
        ] += len(eligibility)

        totals[
            "document_candidates"
        ] += len(documents)

        totals[
            "process_candidates"
        ] += len(process)

        totals[
            "linked_import_rows"
        ] += len(rows)

        service_plans.append(
            {
                "database_pk":
                    service.pk,

                "service_id":
                    service.service_id,

                "title":
                    service.title,

                "domain":
                    (
                        str(service.domain)
                        if service.domain
                        else ""
                    ),

                "category":
                    (
                        str(service.category)
                        if service.category
                        else ""
                    ),

                "service_kind":
                    service.service_kind,

                "structured_fields":
                    (
                        service_structured_snapshot(
                            service
                        )
                    ),

                "existing_counts": {
                    "eligibility_rules":
                        service
                        .eligibility_rules
                        .count(),

                    "documents":
                        service
                        .document_requirements
                        .count(),

                    "process_steps":
                        service
                        .process_steps
                        .count(),

                    "content_sections":
                        service
                        .content_sections
                        .count(),

                    "sources":
                        service
                        .sources
                        .count(),

                    "classifications":
                        service
                        .classification_links
                        .count(),
                },

                "candidate_sections":
                    sections,

                "candidate_eligibility":
                    eligibility,

                "candidate_documents":
                    documents,

                "candidate_process":
                    process,

                "completeness":
                    completeness,

                "evidence":
                    evidence,
            }
        )

    report = {
        "mode":
            "READ_ONLY",

        "database_writes":
            False,

        "generated_at":
            datetime.now()
            .isoformat(),

        "workbook":
            census,

        "database": {
            "published_services":
                services.count(),

            "import_rows":
                ImportRow.objects.count(),

            "eligibility_rules":
                EligibilityRule
                .objects.count(),

            "document_requirements":
                DocumentRequirement
                .objects.count(),

            "process_steps":
                ProcessStep
                .objects.count(),

            "service_content_sections":
                ServiceContentSection
                .objects.count(),
        },

        "knowledge_models":
            knowledge_models,

        "candidate_totals":
            dict(totals),

        "missing_counts":
            dict(missing_counts),

        "services":
            service_plans,
    }

    return report


def save_plan():

    plan = build_intelligence_plan()

    stamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output = (
        AUDIT_DIR
        / (
            "intelligent_ingestion_"
            f"baseline_{stamp}.json"
        )
    )

    output.write_text(
        json.dumps(
            json_safe(plan),
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return plan, output
