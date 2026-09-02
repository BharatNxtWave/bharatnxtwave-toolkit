"""
BharatNXT Wave supporting-data transformation planner.

No database writes.

Transforms approved staged data into in-memory plans for:
- ServiceSource
- ReferenceItem
- ComparisonMatrix / ComparisonEntry
- ServiceContentSection
- CommunicationTemplate

Depends on the already-verified Step 11A core engine.
"""

import hashlib
import re
import statistics

from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from toolkit.importing.transformation import (
    EXPECTED_SOURCE_SHA256,
    build_core_transformation_plan,
    cells,
    column_value,
    fields,
    load_verified_context,
    normalize_identity,
    payload_value,
)

from toolkit.models import (
    CommunicationTemplate,
    ComparisonEntry,
    ComparisonMatrix,
    ReferenceItem,
    ServiceContentSection,
    ServiceSource,
)


EXPECTED_BENEFIT_ROWS = 44
EXPECTED_KNOWLEDGE_ROWS = 207
EXPECTED_KNOWLEDGE_SHEETS = 5
EXPECTED_ONBOARDING_ROWS = 20
EXPECTED_LOAN_DATA_ROWS = 21

KNOWLEDGE_SHEETS = (
    "START_UP INDIA",
    "TAX_CERT",
    "SEED FUND",
    "PVT",
    "LLP",
)

TIE_BREAK_SHEETS = {
    "SEED FUND",
    "LLP",
}

AMOUNT_REFERENCE_COLUMNS = {
    "G": "AMOUNT_DEDUCTIONS_STANDARDS",
    "H": "AMOUNT_DEDUCTIONS_SHORT_FULL_NAME",
    "I": "AMOUNT_DEDUCTIONS_INDUSTRIES",
    "J": "AMOUNT_DEDUCTIONS_IAF_IAS",
    "K": "AMOUNT_DEDUCTIONS_NON_IAF",
}

SOURCE_FIELDS = {
    "portal_link": {
        "source_kind": "APPLICATION",
        "source_name": "Application / Portal",
    },
    "flyer": {
        "source_kind": "FLYER",
        "source_name": "Flyer",
    },
    "additional_info": {
        "source_kind": "REFERENCE",
        "source_name": "Additional Reference",
    },
}


# ============================================================
# GENERIC HELPERS
# ============================================================

def sha256_file(path):
    digest = hashlib.sha256()

    with Path(path).open("rb") as handle:
        for chunk in iter(
            lambda: handle.read(1024 * 1024),
            b"",
        ):
            digest.update(chunk)

    return digest.hexdigest()


def find_verified_workbook():
    source_dir = Path(
        "confidential_source"
    )

    matches = []

    for candidate in source_dir.glob(
        "*.xlsx"
    ):
        if candidate.name.startswith(
            "~$"
        ):
            continue

        try:
            if (
                sha256_file(candidate)
                == EXPECTED_SOURCE_SHA256
            ):
                matches.append(
                    candidate
                )
        except OSError:
            continue

    if len(matches) != 1:
        raise RuntimeError(
            "Expected exactly one hash-verified "
            "source workbook."
        )

    return matches[0]


def sorted_cell_items(row):
    items = list(
        cells(row).items()
    )

    def key(item):
        try:
            return column_index_from_string(
                item[0]
            )
        except Exception:
            return 999999

    return sorted(
        items,
        key=key,
    )


def row_text(row):
    parts = []

    for column, payload in (
        sorted_cell_items(row)
    ):
        value = payload_value(
            payload
        )

        if value:
            parts.append(
                value
            )

    return "\n".join(
        parts
    ).strip()


def valid_http_url(value):
    if not value:
        return False

    try:
        parsed = urlparse(
            str(value).strip()
        )
    except Exception:
        return False

    return (
        parsed.scheme
        in {
            "http",
            "https",
        }
        and bool(
            parsed.netloc
        )
    )


def extract_urls(payload):
    """
    Return only verified http/https URLs.

    No scheme guessing.
    No www -> https invention.
    """

    if not isinstance(
        payload,
        dict,
    ):
        return []

    found = []

    def add(
        candidate,
        method,
    ):
        candidate = str(
            candidate or ""
        ).strip()

        if not candidate:
            return

        if not valid_http_url(
            candidate
        ):
            return

        pair = (
            candidate,
            method,
        )

        if pair not in found:
            found.append(
                pair
            )

    add(
        payload.get(
            "hyperlink_target"
        ),
        "HYPERLINK_TARGET",
    )

    add(
        payload.get(
            "hyperlink_location"
        ),
        "HYPERLINK_LOCATION",
    )

    value = payload_value(
        payload
    )

    if value:
        formula_match = re.search(
            r'HYPERLINK\s*\(\s*"([^"]+)"',
            value,
            flags=re.I,
        )

        if formula_match:
            add(
                formula_match.group(1),
                "FORMULA_HYPERLINK",
            )

        for match in re.findall(
            r'https?://[^\s<>"\']+',
            value,
            flags=re.I,
        ):
            add(
                match.rstrip(
                    ".,);]"
                ),
                "PLAIN_URL",
            )

    return found


def validate_length(
    model,
    field_name,
    value,
):
    field = model._meta.get_field(
        field_name
    )

    max_length = getattr(
        field,
        "max_length",
        None,
    )

    if (
        max_length
        and value is not None
        and len(str(value)) > max_length
    ):
        raise RuntimeError(
            f"{model.__name__}.{field_name} "
            "exceeds max_length."
        )


# ============================================================
# SERVICE SOURCES
# ============================================================

def build_service_sources(
    core_plan,
    staged_rows,
):
    services = core_plan[
        "services"
    ]

    structured_by_row_id = {}

    for identity, plan in (
        services.items()
    ):
        if (
            plan["source_family"]
            != "STRUCTURED"
        ):
            continue

        for row_id in (
            plan["source_row_ids"]
        ):
            structured_by_row_id[
                row_id
            ] = identity

    plans = []

    unresolved_portal_or_flyer = 0

    additional_text_only = 0

    seen = set()

    for row in staged_rows:
        identity = (
            structured_by_row_id.get(
                row.id
            )
        )

        if not identity:
            continue

        for field_name, config in (
            SOURCE_FIELDS.items()
        ):
            payload = fields(
                row
            ).get(
                field_name,
                {},
            )

            value = payload_value(
                payload
            )

            urls = extract_urls(
                payload
            )

            if (
                field_name
                in {
                    "portal_link",
                    "flyer",
                }
                and value
                and not urls
            ):
                unresolved_portal_or_flyer += 1

            if (
                field_name
                == "additional_info"
                and value
                and not urls
            ):
                additional_text_only += 1

            for url, method in urls:
                key = (
                    identity,
                    row.id,
                    config[
                        "source_kind"
                    ],
                    url,
                )

                if key in seen:
                    continue

                seen.add(
                    key
                )

                source_reference = (
                    f"{row.sheet_name}:"
                    f"{row.source_row_number}:"
                    f"{field_name}"
                )

                plans.append(
                    {
                        "service_identity":
                            identity,
                        "source_name":
                            config[
                                "source_name"
                            ],
                        "source_url":
                            url,
                        "source_kind":
                            config[
                                "source_kind"
                            ],
                        "import_row_id":
                            row.id,
                        "source_reference":
                            source_reference,
                        "is_official":
                            False,
                        "notes":
                            "",
                        "extraction_method":
                            method,
                    }
                )

    source_kind_choices = {
        str(value)
        for value, label
        in ServiceSource._meta.get_field(
            "source_kind"
        ).choices
    }

    for item in plans:
        if (
            item[
                "service_identity"
            ]
            not in services
        ):
            raise RuntimeError(
                "ServiceSource references "
                "unknown Service."
            )

        if (
            item[
                "source_kind"
            ]
            not in source_kind_choices
        ):
            raise RuntimeError(
                "Invalid ServiceSource kind."
            )

        if not valid_http_url(
            item["source_url"]
        ):
            raise RuntimeError(
                "Invalid planned ServiceSource URL."
            )

        validate_length(
            ServiceSource,
            "source_name",
            item["source_name"],
        )

        validate_length(
            ServiceSource,
            "source_url",
            item["source_url"],
        )

        validate_length(
            ServiceSource,
            "source_reference",
            item[
                "source_reference"
            ],
        )

    return {
        "plans": plans,
        "unresolved_portal_or_flyer":
            unresolved_portal_or_flyer,
        "additional_text_only":
            additional_text_only,
    }


# ============================================================
# REFERENCE ITEMS
# ============================================================

def build_reference_items(
    staged_rows,
):
    plans = []

    benefit_rows = [
        row
        for row in staged_rows
        if (
            row.sheet_name
            == "BENEFITS"
            and row.raw_data.get(
                "_meta",
                {},
            ).get(
                "row_role"
            )
            == "DATA"
        )
    ]

    if (
        len(benefit_rows)
        != EXPECTED_BENEFIT_ROWS
    ):
        raise RuntimeError(
            "Expected 44 BENEFITS data rows."
        )

    for row in benefit_rows:
        key = column_value(
            row,
            "A",
        )

        value = column_value(
            row,
            "B",
        )

        if not key or not value:
            raise RuntimeError(
                "BENEFITS data row missing "
                "key or value."
            )

        plans.append(
            {
                "dataset_name":
                    "BENEFITS",
                "key":
                    key,
                "value":
                    value,
                "metadata": {
                    "source_sheet":
                        "BENEFITS",
                    "source_column":
                        "B",
                },
                "visibility":
                    "BDE",
                "source_import_row_id":
                    row.id,
                "source_family":
                    "BENEFITS",
            }
        )

    amount_rows = [
        row
        for row in staged_rows
        if (
            row.sheet_name
            == "AMOUNT DEDUCTIONS"
            and row.raw_data.get(
                "_meta",
                {},
            ).get(
                "row_role"
            )
            == "DATA"
        )
    ]

    for row in amount_rows:
        for column, dataset_name in (
            AMOUNT_REFERENCE_COLUMNS.items()
        ):
            value = column_value(
                row,
                column,
            )

            if not value:
                continue

            plans.append(
                {
                    "dataset_name":
                        dataset_name,
                    "key":
                        (
                            f"ROW_"
                            f"{row.source_row_number}"
                        ),
                    "value":
                        value,
                    "metadata": {
                        "source_sheet":
                            "AMOUNT DEDUCTIONS",
                        "source_column":
                            column,
                    },
                    "visibility":
                        "ADMIN_ONLY",
                    "source_import_row_id":
                        row.id,
                    "source_family":
                        "AMOUNT_DEDUCTIONS_REFERENCE",
                }
            )

    visibility_choices = {
        str(value)
        for value, label
        in ReferenceItem._meta.get_field(
            "visibility"
        ).choices
    }

    for item in plans:
        if (
            item["visibility"]
            not in visibility_choices
        ):
            raise RuntimeError(
                "Invalid ReferenceItem visibility."
            )

        if not item["value"]:
            raise RuntimeError(
                "ReferenceItem has blank value."
            )

        validate_length(
            ReferenceItem,
            "dataset_name",
            item[
                "dataset_name"
            ],
        )

        validate_length(
            ReferenceItem,
            "key",
            item["key"],
        )

    return plans


# ============================================================
# LOAN COMPARISON MATRIX
# ============================================================

def build_comparison_plan(
    staged_rows,
):
    loan_rows = [
        row
        for row in staged_rows
        if row.sheet_name
        == "LOAN"
    ]

    title_rows = [
        row
        for row in loan_rows
        if row.raw_data.get(
            "_meta",
            {},
        ).get(
            "row_role"
        )
        == "TITLE"
    ]

    header_rows = [
        row
        for row in loan_rows
        if row.raw_data.get(
            "_meta",
            {},
        ).get(
            "row_role"
        )
        == "HEADER"
    ]

    data_rows = [
        row
        for row in loan_rows
        if row.raw_data.get(
            "_meta",
            {},
        ).get(
            "row_role"
        )
        == "DATA"
    ]

    if (
        len(title_rows) != 1
        or len(header_rows) != 1
        or len(data_rows)
        != EXPECTED_LOAN_DATA_ROWS
    ):
        raise RuntimeError(
            "Unexpected LOAN matrix structure."
        )

    matrix_name = row_text(
        title_rows[0]
    )

    if not matrix_name:
        raise RuntimeError(
            "LOAN matrix title is blank."
        )

    validate_length(
        ComparisonMatrix,
        "name",
        matrix_name,
    )

    validate_length(
        ComparisonMatrix,
        "source_sheet",
        "LOAN",
    )

    headers = {}

    for column, payload in (
        sorted_cell_items(
            header_rows[0]
        )
    ):
        value = payload_value(
            payload
        )

        if value:
            headers[
                column
            ] = value

    entries = []

    missing_header_cells = 0

    for row in data_rows:
        row_label = column_value(
            row,
            "A",
        )

        for column, payload in (
            sorted_cell_items(row)
        ):
            value = payload_value(
                payload
            )

            if not value:
                continue

            column_name = headers.get(
                column,
                "",
            )

            if not column_name:
                missing_header_cells += 1
                continue

            validate_length(
                ComparisonEntry,
                "column_name",
                column_name,
            )

            validate_length(
                ComparisonEntry,
                "row_label",
                row_label,
            )

            entries.append(
                {
                    "row_number":
                        row.source_row_number,
                    "column_name":
                        column_name,
                    "row_label":
                        row_label,
                    "value_raw":
                        value,
                    "service_identity":
                        None,
                    "source_import_row_id":
                        row.id,
                }
            )

    if missing_header_cells:
        raise RuntimeError(
            "LOAN data contains nonblank "
            "cells without a source header."
        )

    if not entries:
        raise RuntimeError(
            "LOAN comparison produced zero entries."
        )

    return {
        "matrix": {
            "name":
                matrix_name,
            "source_sheet":
                "LOAN",
            "import_batch_id":
                5,
            "metadata": {
                "title_source_row":
                    title_rows[
                        0
                    ].source_row_number,
                "header_source_row":
                    header_rows[
                        0
                    ].source_row_number,
                "data_row_count":
                    len(
                        data_rows
                    ),
            },
        },
        "entries":
            entries,
    }


# ============================================================
# KNOWLEDGE ANCHOR RESOLUTION
# ============================================================

def service_match_score(
    candidate,
    text,
):
    candidate = normalize_identity(
        candidate
    )

    text = normalize_identity(
        text
    )

    if not candidate or not text:
        return 0.0

    if candidate == text:
        return 1.0

    candidate_tokens = set(
        candidate.split()
    )

    text_tokens = set(
        text.split()
    )

    if not candidate_tokens:
        return 0.0

    shared = (
        candidate_tokens
        & text_tokens
    )

    coverage = (
        len(shared)
        / len(
            candidate_tokens
        )
    )

    union = (
        candidate_tokens
        | text_tokens
    )

    jaccard = (
        len(shared)
        / len(union)
        if union
        else 0.0
    )

    sequence = SequenceMatcher(
        None,
        candidate,
        text,
    ).ratio()

    if (
        len(candidate_tokens) >= 2
        and candidate_tokens.issubset(
            text_tokens
        )
    ):
        return max(
            0.96,
            sequence,
        )

    token_score = (
        0.65 * coverage
        + 0.35 * jaccard
    )

    return min(
        1.0,
        max(
            token_score,
            sequence * 0.80,
        ),
    )


def content_evidence_tie_break(
    sheet_name,
    sheet_rows,
    candidate_identities,
):
    content_entries = []

    for row in sheet_rows:
        for column, payload in (
            sorted_cell_items(row)
        ):
            value = payload_value(
                payload
            )

            if value:
                content_entries.append(
                    (
                        row.source_row_number,
                        value,
                    )
                )

    sheet_tokens = set(
        normalize_identity(
            sheet_name
        ).split()
    )

    evidence = []

    for identity in (
        candidate_identities
    ):
        candidate_tokens = set(
            identity.split()
        )

        if not candidate_tokens:
            continue

        exact = 0
        containment = 0
        earliest = None

        for row_number, text in (
            content_entries
        ):
            normalized = (
                normalize_identity(
                    text
                )
            )

            text_tokens = set(
                normalized.split()
            )

            if normalized == identity:
                exact += 1

            if (
                candidate_tokens
                and candidate_tokens.issubset(
                    text_tokens
                )
            ):
                containment += 1

                if (
                    earliest is None
                    or row_number
                    < earliest
                ):
                    earliest = row_number

        evidence.append(
            {
                "identity":
                    identity,
                "exact":
                    exact,
                "containment":
                    containment,
                "sheet_overlap":
                    len(
                        candidate_tokens
                        & sheet_tokens
                    ),
                "earliest":
                    (
                        earliest
                        if earliest
                        is not None
                        else 999999
                    ),
            }
        )

    evidence.sort(
        key=lambda item: (
            item["exact"],
            item["containment"],
            item[
                "sheet_overlap"
            ],
            -item["earliest"],
        ),
        reverse=True,
    )

    if len(evidence) < 2:
        raise RuntimeError(
            "Insufficient anchor candidates."
        )

    top = evidence[0]
    second = evidence[1]

    top_tuple = (
        top["exact"],
        top["containment"],
        top[
            "sheet_overlap"
        ],
        -top["earliest"],
    )

    second_tuple = (
        second["exact"],
        second["containment"],
        second[
            "sheet_overlap"
        ],
        -second["earliest"],
    )

    if not (
        top_tuple > second_tuple
        and (
            top["exact"] >= 1
            or top[
                "containment"
            ] >= 2
        )
    ):
        raise RuntimeError(
            f"{sheet_name} knowledge anchor "
            "remains unresolved."
        )

    return top[
        "identity"
    ]


def resolve_knowledge_anchors(
    core_plan,
    staged_rows,
):
    services = core_plan[
        "services"
    ]

    all_identities = set(
        services
    )

    commercial_identities = {
        identity
        for identity, plan
        in services.items()
        if plan[
            "source_family"
        ]
        == "COMMERCIAL"
    }

    results = {}

    for sheet_name in (
        KNOWLEDGE_SHEETS
    ):
        sheet_rows = [
            row
            for row in staged_rows
            if row.sheet_name
            == sheet_name
        ]

        texts = [
            sheet_name
        ]

        for row in sheet_rows:
            for column, payload in (
                sorted_cell_items(row)
            ):
                value = payload_value(
                    payload
                )

                if value:
                    texts.append(
                        value
                    )

        scores = []

        for identity in (
            all_identities
        ):
            best = max(
                (
                    service_match_score(
                        identity,
                        text,
                    )
                    for text in texts
                ),
                default=0.0,
            )

            scores.append(
                (
                    best,
                    identity,
                )
            )

        scores.sort(
            reverse=True,
            key=lambda item: item[0],
        )

        top_score = scores[
            0
        ][0]

        second_score = scores[
            1
        ][0]

        gap = (
            top_score
            - second_score
        )

        if (
            sheet_name
            not in TIE_BREAK_SHEETS
        ):
            if not (
                top_score >= 0.95
                and gap >= 0.05
            ):
                raise RuntimeError(
                    f"{sheet_name} no longer "
                    "has the approved high-confidence "
                    "knowledge anchor."
                )

            identity = scores[
                0
            ][1]

            method = (
                "FUZZY_HIGH_CONFIDENCE"
            )

        else:
            identity = (
                content_evidence_tie_break(
                    sheet_name,
                    sheet_rows,
                    commercial_identities,
                )
            )

            method = (
                "CONTENT_EVIDENCE"
            )

        if (
            identity
            not in commercial_identities
        ):
            raise RuntimeError(
                f"{sheet_name} anchor is no longer "
                "a commercial Service as previously audited."
            )

        results[
            sheet_name
        ] = {
            "service_identity":
                identity,
            "method":
                method,
            "row_count":
                len(
                    sheet_rows
                ),
        }

    method_counts = Counter(
        result[
            "method"
        ]
        for result in (
            results.values()
        )
    )

    if (
        method_counts[
            "FUZZY_HIGH_CONFIDENCE"
        ]
        != 3
        or method_counts[
            "CONTENT_EVIDENCE"
        ]
        != 2
    ):
        raise RuntimeError(
            "Knowledge anchor method counts "
            "differ from approved audits."
        )

    return results


# ============================================================
# KNOWLEDGE STYLE METADATA
# ============================================================

def knowledge_heading_rows(
    workbook_path,
    staged_rows,
):
    workbook = load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )

    result = {}

    try:
        for sheet_name in (
            KNOWLEDGE_SHEETS
        ):
            worksheet = workbook[
                sheet_name
            ]

            relevant_rows = [
                row
                for row in staged_rows
                if row.sheet_name
                == sheet_name
            ]

            source_row_numbers = {
                row.source_row_number
                for row in relevant_rows
            }

            font_sizes = []

            for row_number in (
                source_row_numbers
            ):
                for cell in worksheet[
                    row_number
                ]:
                    if (
                        cell.value
                        is None
                    ):
                        continue

                    if (
                        cell.font.size
                        is not None
                    ):
                        try:
                            font_sizes.append(
                                float(
                                    cell.font.size
                                )
                            )
                        except (
                            TypeError,
                            ValueError,
                        ):
                            pass

            median_font = (
                statistics.median(
                    font_sizes
                )
                if font_sizes
                else 11.0
            )

            merged_rows = set()

            for merged_range in (
                worksheet.merged_cells.ranges
            ):
                for row_number in range(
                    merged_range.min_row,
                    merged_range.max_row + 1,
                ):
                    merged_rows.add(
                        row_number
                    )

            headings = set()

            for row_number in sorted(
                source_row_numbers
            ):
                nonempty = []

                row_bold = False

                row_large = False

                for cell in worksheet[
                    row_number
                ]:
                    if (
                        cell.value
                        is None
                    ):
                        continue

                    nonempty.append(
                        cell
                    )

                    if cell.font.bold:
                        row_bold = True

                    if (
                        cell.font.size
                        is not None
                    ):
                        try:
                            if (
                                float(
                                    cell.font.size
                                )
                                > median_font
                            ):
                                row_large = True
                        except (
                            TypeError,
                            ValueError,
                        ):
                            pass

                if (
                    len(nonempty) == 1
                    and (
                        row_bold
                        or row_large
                        or row_number
                        in merged_rows
                    )
                ):
                    headings.add(
                        row_number
                    )

            result[
                sheet_name
            ] = headings

    finally:
        workbook.close()

    return result


# ============================================================
# KNOWLEDGE SECTION TYPE
# ============================================================

SECTION_RULES = (
    (
        "BENEFITS",
        (
            "benefit",
            "advantage",
        ),
    ),
    (
        "ELIGIBILITY",
        (
            "eligibility",
            "eligible",
        ),
    ),
    (
        "FUNDING",
        (
            "fund",
            "funding",
            "finance",
        ),
    ),
    (
        "PROCESS",
        (
            "process",
            "procedure",
            "steps",
            "how to",
        ),
    ),
    (
        "DOCUMENTS",
        (
            "document",
            "documents",
        ),
    ),
    (
        "TIMELINE",
        (
            "timeline",
            "time line",
            "duration",
        ),
    ),
    (
        "SCOPE",
        (
            "scope",
            "coverage",
        ),
    ),
    (
        "GLOSSARY",
        (
            "glossary",
            "definition",
        ),
    ),
    (
        "COMMERCIAL",
        (
            "commercial",
            "charges",
            "pricing",
            "fees",
        ),
    ),
    (
        "NOTES",
        (
            "note",
            "notes",
            "important",
        ),
    ),
    (
        "OVERVIEW",
        (
            "overview",
            "about",
            "introduction",
        ),
    ),
)


def infer_section_type(
    heading,
    allowed,
):
    normalized = (
        normalize_identity(
            heading
        )
    )

    for section_type, terms in (
        SECTION_RULES
    ):
        for term in terms:
            term = normalize_identity(
                term
            )

            if (
                term
                and re.search(
                    r"(?:^|\s)"
                    + re.escape(
                        term
                    )
                    + r"(?:\s|$)",
                    normalized,
                )
            ):
                if section_type in allowed:
                    return section_type

    return (
        "OTHER"
        if "OTHER" in allowed
        else next(
            iter(
                allowed
            )
        )
    )


# ============================================================
# KNOWLEDGE CONTENT SECTIONS
# ============================================================

def build_knowledge_sections(
    core_plan,
    staged_rows,
):
    workbook_path = (
        find_verified_workbook()
    )

    anchors = (
        resolve_knowledge_anchors(
            core_plan,
            staged_rows,
        )
    )

    heading_rows = (
        knowledge_heading_rows(
            workbook_path,
            staged_rows,
        )
    )

    section_type_choices = {
        str(value)
        for value, label
        in ServiceContentSection._meta.get_field(
            "section_type"
        ).choices
    }

    visibility_choices = {
        str(value)
        for value, label
        in ServiceContentSection._meta.get_field(
            "visibility"
        ).choices
    }

    if (
        "BDE"
        not in visibility_choices
    ):
        raise RuntimeError(
            "BDE knowledge visibility unavailable."
        )

    plans = []

    covered_rows = set()

    heading_count = 0

    max_title_length = (
        ServiceContentSection._meta.get_field(
            "title"
        ).max_length
    )

    for sheet_name in (
        KNOWLEDGE_SHEETS
    ):
        sheet_rows = sorted(
            (
                row
                for row in staged_rows
                if row.sheet_name
                == sheet_name
            ),
            key=lambda row: (
                row.source_row_number,
                row.id,
            ),
        )

        anchor_identity = (
            anchors[
                sheet_name
            ][
                "service_identity"
            ]
        )

        current_type = "OTHER"

        display_order = 0

        for row in sheet_rows:
            text = row_text(
                row
            )

            if not text:
                raise RuntimeError(
                    "Knowledge ImportRow unexpectedly "
                    "contains no staged text."
                )

            is_heading = (
                row.source_row_number
                in heading_rows[
                    sheet_name
                ]
            )

            title = ""

            if is_heading:
                heading_count += 1

                current_type = (
                    infer_section_type(
                        text,
                        section_type_choices,
                    )
                )

                if (
                    not max_title_length
                    or len(text)
                    <= max_title_length
                ):
                    title = text

            display_order += 1

            plans.append(
                {
                    "service_identity":
                        anchor_identity,
                    "section_type":
                        current_type,
                    "title":
                        title,
                    "content":
                        text,
                    "display_order":
                        display_order,
                    "visibility":
                        "BDE",
                    "source_import_row_id":
                        row.id,
                    "source_sheet":
                        sheet_name,
                    "is_heading":
                        is_heading,
                }
            )

            covered_rows.add(
                row.id
            )

    if (
        len(plans)
        != EXPECTED_KNOWLEDGE_ROWS
    ):
        raise RuntimeError(
            "Expected 207 knowledge section plans."
        )

    if (
        len(covered_rows)
        != EXPECTED_KNOWLEDGE_ROWS
    ):
        raise RuntimeError(
            "Knowledge source-row lineage coverage "
            "is incomplete."
        )

    for item in plans:
        if (
            item[
                "service_identity"
            ]
            not in core_plan[
                "services"
            ]
        ):
            raise RuntimeError(
                "Knowledge section references "
                "unknown Service."
            )

        if (
            item[
                "section_type"
            ]
            not in section_type_choices
        ):
            raise RuntimeError(
                "Invalid knowledge section_type."
            )

        if (
            item[
                "visibility"
            ]
            not in visibility_choices
        ):
            raise RuntimeError(
                "Invalid knowledge visibility."
            )

        if not item[
            "content"
        ]:
            raise RuntimeError(
                "Knowledge section has blank content."
            )

        validate_length(
            ServiceContentSection,
            "title",
            item["title"],
        )

    return {
        "plans":
            plans,
        "anchors":
            anchors,
        "heading_count":
            heading_count,
        "workbook_hash_verified":
            True,
    }


# ============================================================
# ONBOARDING MAIL
# ============================================================

def build_communication_plan(
    staged_rows,
):
    mail_rows = sorted(
        (
            row
            for row in staged_rows
            if row.sheet_name
            == "ONBOARDING_MAIL"
        ),
        key=lambda row: (
            row.source_row_number,
            row.id,
        ),
    )

    if (
        len(mail_rows)
        != EXPECTED_ONBOARDING_ROWS
    ):
        raise RuntimeError(
            "Expected 20 ONBOARDING_MAIL rows."
        )

    subject_rows = []

    texts_by_row = {}

    for row in mail_rows:
        text = row_text(
            row
        )

        texts_by_row[
            row.source_row_number
        ] = text

        match = re.match(
            r"^\s*subject\s*[:\-]\s*(.*)$",
            text,
            flags=re.I,
        )

        if match:
            subject_rows.append(
                (
                    row,
                    match.group(1).strip(),
                )
            )

    if len(subject_rows) != 1:
        raise RuntimeError(
            "Expected exactly one onboarding "
            "mail Subject line."
        )

    subject_row, subject = (
        subject_rows[0]
    )

    minimum_row = min(
        texts_by_row
    )

    maximum_row = max(
        texts_by_row
    )

    body_lines = []

    for row_number in range(
        minimum_row,
        maximum_row + 1,
    ):
        if (
            row_number
            == subject_row.source_row_number
        ):
            continue

        body_lines.append(
            texts_by_row.get(
                row_number,
                "",
            )
        )

    body = "\n".join(
        body_lines
    ).strip()

    if not body:
        raise RuntimeError(
            "Onboarding email body is blank."
        )

    plan = {
        "template_key":
            "bharatnxt-onboarding-mail-v1",
        "title":
            "BharatNXT Onboarding Mail",
        "subject":
            subject,
        "body":
            body,
        "status":
            "DRAFT",
        "visibility":
            "ADMIN_ONLY",
        "source_import_batch_id":
            5,
        "source_sheet":
            "ONBOARDING_MAIL",
    }

    status_choices = {
        str(value)
        for value, label
        in CommunicationTemplate._meta.get_field(
            "status"
        ).choices
    }

    visibility_choices = {
        str(value)
        for value, label
        in CommunicationTemplate._meta.get_field(
            "visibility"
        ).choices
    }

    if (
        plan["status"]
        not in status_choices
    ):
        raise RuntimeError(
            "Invalid communication status."
        )

    if (
        plan["visibility"]
        not in visibility_choices
    ):
        raise RuntimeError(
            "Invalid communication visibility."
        )

    for field_name in (
        "template_key",
        "title",
        "subject",
        "source_sheet",
    ):
        validate_length(
            CommunicationTemplate,
            field_name,
            plan[
                field_name
            ],
        )

    return plan


# ============================================================
# PUBLIC WHOLE-SUPPORTING PLAN
# ============================================================

def build_supporting_transformation_plan(
    core_plan=None,
):
    (
        contract,
        batch,
        staged_rows,
    ) = load_verified_context()

    if core_plan is None:
        core_plan = (
            build_core_transformation_plan()
        )

    if (
        core_plan[
            "contract_version"
        ]
        != 3
    ):
        raise RuntimeError(
            "Supporting planner requires "
            "Contract v3 core plan."
        )

    sources = build_service_sources(
        core_plan,
        staged_rows,
    )

    references = (
        build_reference_items(
            staged_rows
        )
    )

    comparison = (
        build_comparison_plan(
            staged_rows
        )
    )

    knowledge = (
        build_knowledge_sections(
            core_plan,
            staged_rows,
        )
    )

    communication = (
        build_communication_plan(
            staged_rows
        )
    )

    return {
        "contract_version":
            contract[
                "contract"
            ][
                "version"
            ],
        "batch_id":
            batch.id,
        "sources":
            sources,
        "references":
            references,
        "comparison":
            comparison,
        "knowledge":
            knowledge,
        "communication":
            communication,
    }
