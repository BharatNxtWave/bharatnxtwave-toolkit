from collections import Counter
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .sheet_registry import (
    EXPECTED_SHEETS,
    SCHEME_TABLES,
    canonical_header,
    sheet_family,
)


MAX_HEADER_SCAN_ROWS = 15
MAX_INSPECT_COLUMNS = 30


def file_sha256(path):

    digest = sha256()

    with Path(path).open("rb") as file:

        for chunk in iter(
            lambda: file.read(1024 * 1024),
            b"",
        ):
            digest.update(chunk)

    return digest.hexdigest()


def normalized_text(value):

    if value is None:
        return ""

    if isinstance(
        value,
        (datetime, date),
    ):
        return value.isoformat()

    return re.sub(
        r"\s+",
        " ",
        str(value),
    ).strip()


def normalized_identity(value):

    text = normalized_text(
        value
    ).casefold()

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def build_merged_lookup(worksheet):

    lookup = {}

    for merged_range in (
        worksheet.merged_cells.ranges
    ):

        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row

        top_left = (
            min_row,
            min_col,
        )

        for row in range(
            min_row,
            max_row + 1,
        ):

            for column in range(
                min_col,
                max_col + 1,
            ):

                lookup[
                    (row, column)
                ] = top_left

    return lookup


def resolved_cell(
    worksheet,
    row,
    column,
    merged_lookup,
):

    cell = worksheet.cell(
        row=row,
        column=column,
    )

    if isinstance(
        cell,
        MergedCell,
    ):

        top_left = merged_lookup.get(
            (row, column)
        )

        if top_left:

            return worksheet.cell(
                row=top_left[0],
                column=top_left[1],
            )

    return cell


def cell_has_content(cell):

    if cell is None:
        return False

    if normalized_text(
        cell.value
    ):
        return True

    hyperlink = getattr(
        cell,
        "hyperlink",
        None,
    )

    if hyperlink and getattr(
        hyperlink,
        "target",
        None,
    ):
        return True

    return False


def find_scheme_header(
    worksheet,
    merged_lookup,
):

    best = None

    max_column = min(
        worksheet.max_column or 0,
        MAX_INSPECT_COLUMNS,
    )

    max_row = min(
        worksheet.max_row or 0,
        MAX_HEADER_SCAN_ROWS,
    )

    for row_number in range(
        1,
        max_row + 1,
    ):

        mapping = {}

        for column in range(
            1,
            max_column + 1,
        ):

            cell = resolved_cell(
                worksheet,
                row_number,
                column,
                merged_lookup,
            )

            canonical = canonical_header(
                cell.value
            )

            if (
                canonical
                and canonical not in mapping
            ):
                mapping[
                    canonical
                ] = column

        score = len(mapping)

        # A real scheme-table header must contain
        # Scheme Name. Other fields may vary.
        if "scheme_name" not in mapping:
            continue

        candidate = (
            score,
            -row_number,
            row_number,
            mapping,
        )

        if (
            best is None
            or candidate > best
        ):
            best = candidate

    if best is None:
        return None, {}

    return (
        best[2],
        best[3],
    )


def inspect_scheme_sheet(
    worksheet,
):

    merged_lookup = build_merged_lookup(
        worksheet
    )

    header_row, mapping = (
        find_scheme_header(
            worksheet,
            merged_lookup,
        )
    )

    result = {
        "header_row": header_row,
        "fields": sorted(mapping),
        "active_rows": 0,
        "rows_missing_scheme_name": 0,
        "portal_hyperlink_cells": 0,
        "merged_ranges": len(
            worksheet.merged_cells.ranges
        ),
        "_identities": [],
    }

    if header_row is None:
        return result

    meaningful_fields = {
        key: column
        for key, column
        in mapping.items()
        if key != "serial_number"
    }

    for row_number in range(
        header_row + 1,
        (worksheet.max_row or header_row) + 1,
    ):

        resolved = {}

        active = False

        for field, column in (
            meaningful_fields.items()
        ):

            cell = resolved_cell(
                worksheet,
                row_number,
                column,
                merged_lookup,
            )

            resolved[field] = cell

            if cell_has_content(cell):
                active = True

        if not active:
            continue

        result["active_rows"] += 1

        scheme_cell = resolved.get(
            "scheme_name"
        )

        scheme_name = (
            normalized_text(
                scheme_cell.value
            )
            if scheme_cell
            else ""
        )

        if not scheme_name:
            result[
                "rows_missing_scheme_name"
            ] += 1

        identity = normalized_identity(
            scheme_name
        )

        if identity:
            result[
                "_identities"
            ].append(identity)

        portal_cell = resolved.get(
            "portal_link"
        )

        if portal_cell:

            hyperlink = getattr(
                portal_cell,
                "hyperlink",
                None,
            )

            if (
                hyperlink
                and getattr(
                    hyperlink,
                    "target",
                    None,
                )
            ):
                result[
                    "portal_hyperlink_cells"
                ] += 1

    return result


def inspect_nonstandard_sheet(
    worksheet,
):

    merged_lookup = build_merged_lookup(
        worksheet
    )

    meaningful_rows = 0
    populated_cells = 0
    hyperlink_cells = 0

    max_column = min(
        worksheet.max_column or 0,
        MAX_INSPECT_COLUMNS,
    )

    for row_number in range(
        1,
        (worksheet.max_row or 0) + 1,
    ):

        row_has_content = False

        for column in range(
            1,
            max_column + 1,
        ):

            cell = resolved_cell(
                worksheet,
                row_number,
                column,
                merged_lookup,
            )

            if not cell_has_content(
                cell
            ):
                continue

            row_has_content = True
            populated_cells += 1

            hyperlink = getattr(
                cell,
                "hyperlink",
                None,
            )

            if (
                hyperlink
                and getattr(
                    hyperlink,
                    "target",
                    None,
                )
            ):
                hyperlink_cells += 1

        if row_has_content:
            meaningful_rows += 1

    return {
        "meaningful_rows": meaningful_rows,
        "populated_cells": populated_cells,
        "hyperlink_cells": hyperlink_cells,
        "merged_ranges": len(
            worksheet.merged_cells.ranges
        ),
    }


def inspect_workbook(path):

    path = Path(path)

    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )

    result = {
        "sheet_count": len(
            workbook.sheetnames
        ),
        "sheets": [],
        "missing_expected_sheets": [],
        "unexpected_sheets": [],
        "structured_scheme_rows": 0,
        "cross_sheet_duplicate_names": 0,
        "cross_sheet_duplicate_occurrences": 0,
    }

    expected = set(
        EXPECTED_SHEETS
    )

    actual = set(
        workbook.sheetnames
    )

    result[
        "missing_expected_sheets"
    ] = sorted(
        expected - actual
    )

    result[
        "unexpected_sheets"
    ] = sorted(
        actual - expected
    )

    identity_sheets = {}

    try:

        for sheet_name in (
            workbook.sheetnames
        ):

            worksheet = workbook[
                sheet_name
            ]

            family = sheet_family(
                sheet_name
            )

            entry = {
                "name": sheet_name,
                "family": family,
            }

            if sheet_name in SCHEME_TABLES:

                details = (
                    inspect_scheme_sheet(
                        worksheet
                    )
                )

                identities = details.pop(
                    "_identities"
                )

                entry.update(
                    details
                )

                result[
                    "structured_scheme_rows"
                ] += details[
                    "active_rows"
                ]

                for identity in identities:

                    identity_sheets.setdefault(
                        identity,
                        set(),
                    ).add(
                        sheet_name
                    )

            else:

                entry.update(
                    inspect_nonstandard_sheet(
                        worksheet
                    )
                )

            result["sheets"].append(
                entry
            )

    finally:

        workbook.close()

    duplicate_groups = {
        identity: sheets
        for identity, sheets
        in identity_sheets.items()
        if len(sheets) > 1
    }

    result[
        "cross_sheet_duplicate_names"
    ] = len(
        duplicate_groups
    )

    result[
        "cross_sheet_duplicate_occurrences"
    ] = sum(
        len(sheets)
        for sheets
        in duplicate_groups.values()
    )

    return result
