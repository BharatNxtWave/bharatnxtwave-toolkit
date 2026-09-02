from collections import Counter, defaultdict
from datetime import date, datetime, time
from decimal import Decimal
from hashlib import sha256
import json
from pathlib import Path

from django.db import transaction

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from toolkit.models import (
    ImportBatch,
    ImportRow,
    Service,
)

from .sheet_registry import (
    SCHEME_TABLES,
    sheet_family,
)

from .workbook_reader import (
    build_merged_lookup,
    cell_has_content,
    file_sha256,
    find_scheme_header,
    normalized_identity,
    normalized_text,
    resolved_cell,
)


EXPECTED_STRUCTURED_ROWS = 113
EXPECTED_TOTAL_STAGED_ROWS = 478
EXPECTED_DUPLICATE_GROUPS = 9
EXPECTED_DUPLICATE_OCCURRENCES = 21
EXPECTED_SHEET_COUNT = 20

MAX_NONSTANDARD_COLUMNS = 30


def json_safe(value):

    if value is None:
        return ""

    if isinstance(
        value,
        (datetime, date, time),
    ):
        return value.isoformat()

    if isinstance(value, Decimal):
        return str(value)

    if isinstance(
        value,
        (str, int, float, bool),
    ):
        return value

    return str(value)


def make_row_hash(payload):

    encoded = json.dumps(
        payload,
        sort_keys=True,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")

    return sha256(
        encoded
    ).hexdigest()


def cell_payload(
    worksheet,
    row_number,
    column_number,
    merged_lookup,
):

    raw_cell = worksheet.cell(
        row=row_number,
        column=column_number,
    )

    resolved = resolved_cell(
        worksheet,
        row_number,
        column_number,
        merged_lookup,
    )

    payload = {
        "value": json_safe(
            resolved.value
        ),
        "cell_type": getattr(
            resolved,
            "data_type",
            None,
        ),
        "coordinate": (
            resolved.coordinate
        ),
    }

    if isinstance(
        raw_cell,
        MergedCell,
    ):
        origin = merged_lookup.get(
            (
                row_number,
                column_number,
            )
        )

        if origin:
            payload["merged_from"] = (
                f"{get_column_letter(origin[1])}"
                f"{origin[0]}"
            )

    hyperlink = getattr(
        resolved,
        "hyperlink",
        None,
    )

    if hyperlink:

        target = getattr(
            hyperlink,
            "target",
            None,
        )

        location = getattr(
            hyperlink,
            "location",
            None,
        )

        if target:
            payload[
                "hyperlink_target"
            ] = str(target)

        if location:
            payload[
                "hyperlink_location"
            ] = str(location)

    return payload


def payload_has_content(payload):

    if normalized_text(
        payload.get("value")
    ):
        return True

    if payload.get(
        "hyperlink_target"
    ):
        return True

    if payload.get(
        "hyperlink_location"
    ):
        return True

    return False


def nonstandard_row_role(
    sheet_name,
    family,
    row_number,
):

    if sheet_name in {
        "BENEFITS",
        "AMOUNT DEDUCTIONS",
    }:
        if row_number == 1:
            return "HEADER"

        return "DATA"

    if sheet_name == "LOAN":

        if row_number == 1:
            return "TITLE"

        if row_number == 2:
            return "HEADER"

        return "DATA"

    if family == "KNOWLEDGE":
        return "CONTENT"

    if family == "COMMUNICATION":
        return "CONTENT"

    if family == "ROLLING_GRANTS":
        return "DATA"

    return "DATA"


def extract_scheme_rows(
    worksheet,
):

    merged_lookup = (
        build_merged_lookup(
            worksheet
        )
    )

    header_row, mapping = (
        find_scheme_header(
            worksheet,
            merged_lookup,
        )
    )

    if header_row is None:

        raise ValueError(
            f"Scheme header not found: "
            f"{worksheet.title}"
        )

    meaningful_mapping = {
        field: column
        for field, column
        in mapping.items()
        if field != "serial_number"
    }

    rows = []

    for row_number in range(
        header_row + 1,
        (worksheet.max_row or header_row)
        + 1,
    ):

        mapped_fields = {}
        active = False

        for field, column in (
            mapping.items()
        ):

            payload = cell_payload(
                worksheet,
                row_number,
                column,
                merged_lookup,
            )

            mapped_fields[
                field
            ] = payload

            if (
                field
                in meaningful_mapping
                and payload_has_content(
                    payload
                )
            ):
                active = True

        if not active:
            continue

        scheme_payload = (
            mapped_fields.get(
                "scheme_name",
                {},
            )
        )

        scheme_name = (
            normalized_text(
                scheme_payload.get(
                    "value"
                )
            )
        )

        identity = (
            normalized_identity(
                scheme_name
            )
        )

        raw_data = {
            "_meta": {
                "family": (
                    "SCHEME_TABLE"
                ),
                "row_role": "DATA",
                "sheet_name": (
                    worksheet.title
                ),
                "source_row": (
                    row_number
                ),
                "header_row": (
                    header_row
                ),
            },
            "fields": mapped_fields,
        }

        rows.append(
            {
                "sheet_name": (
                    worksheet.title
                ),
                "source_row_number": (
                    row_number
                ),
                "source_key": (
                    identity[:255]
                    if identity
                    else ""
                ),
                "identity": identity,
                "raw_data": raw_data,
                "row_hash": (
                    make_row_hash(
                        raw_data
                    )
                ),
                "validation_status": (
                    "PENDING"
                ),
                "validation_errors": [],
                "validation_warnings": [],
                "candidate_action": (
                    "UNDECIDED"
                ),
            }
        )

    return rows


def extract_nonstandard_rows(
    worksheet,
    family,
):

    merged_lookup = (
        build_merged_lookup(
            worksheet
        )
    )

    max_column = min(
        worksheet.max_column or 0,
        MAX_NONSTANDARD_COLUMNS,
    )

    rows = []

    for row_number in range(
        1,
        (worksheet.max_row or 0)
        + 1,
    ):

        cells = {}
        active = False

        for column_number in range(
            1,
            max_column + 1,
        ):

            payload = cell_payload(
                worksheet,
                row_number,
                column_number,
                merged_lookup,
            )

            if not payload_has_content(
                payload
            ):
                continue

            active = True

            cells[
                get_column_letter(
                    column_number
                )
            ] = payload

        if not active:
            continue

        row_role = (
            nonstandard_row_role(
                worksheet.title,
                family,
                row_number,
            )
        )

        raw_data = {
            "_meta": {
                "family": family,
                "row_role": row_role,
                "sheet_name": (
                    worksheet.title
                ),
                "source_row": (
                    row_number
                ),
            },
            "cells": cells,
        }

        rows.append(
            {
                "sheet_name": (
                    worksheet.title
                ),
                "source_row_number": (
                    row_number
                ),
                "source_key": (
                    f"{worksheet.title}:"
                    f"{row_number}"
                )[:255],
                "identity": "",
                "raw_data": raw_data,
                "row_hash": (
                    make_row_hash(
                        raw_data
                    )
                ),
                "validation_status": (
                    "PENDING"
                ),
                "validation_errors": [],
                "validation_warnings": [],
                "candidate_action": (
                    "UNDECIDED"
                ),
            }
        )

    return rows


def extract_workbook_rows(
    workbook_path,
):

    workbook_path = Path(
        workbook_path
    )

    workbook = load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )

    rows = []
    sheet_counts = Counter()
    family_counts = Counter()

    try:

        if (
            len(workbook.sheetnames)
            != EXPECTED_SHEET_COUNT
        ):
            raise ValueError(
                "Unexpected workbook sheet count: "
                f"{len(workbook.sheetnames)}"
            )

        for sheet_name in (
            workbook.sheetnames
        ):

            worksheet = workbook[
                sheet_name
            ]

            family = sheet_family(
                sheet_name
            )

            if family == "UNKNOWN":
                raise ValueError(
                    "Unknown worksheet encountered: "
                    f"{sheet_name}"
                )

            if sheet_name in (
                SCHEME_TABLES
            ):

                extracted = (
                    extract_scheme_rows(
                        worksheet
                    )
                )

            else:

                extracted = (
                    extract_nonstandard_rows(
                        worksheet,
                        family,
                    )
                )

            rows.extend(
                extracted
            )

            sheet_counts[
                sheet_name
            ] += len(
                extracted
            )

            family_counts[
                family
            ] += len(
                extracted
            )

    finally:

        workbook.close()

    return (
        rows,
        sheet_counts,
        family_counts,
    )


def classify_scheme_candidates(
    rows,
):

    identity_rows = defaultdict(list)
    identity_sheets = defaultdict(set)

    for index, row in enumerate(rows):

        if (
            row["raw_data"]
            ["_meta"]
            ["family"]
            != "SCHEME_TABLE"
        ):
            continue

        identity = row[
            "identity"
        ]

        if not identity:
            continue

        identity_rows[
            identity
        ].append(index)

        identity_sheets[
            identity
        ].add(
            row["sheet_name"]
        )

    cross_sheet_duplicates = {
        identity
        for identity, sheets
        in identity_sheets.items()
        if len(sheets) > 1
    }

    duplicate_occurrences = sum(
        len(
            identity_sheets[
                identity
            ]
        )
        for identity
        in cross_sheet_duplicates
    )

    existing_titles = {
        normalized_identity(title)
        for title
        in Service.objects.values_list(
            "title",
            flat=True,
        )
        if normalized_identity(title)
    }

    for row in rows:

        if (
            row["raw_data"]
            ["_meta"]
            ["family"]
            != "SCHEME_TABLE"
        ):
            continue

        identity = row[
            "identity"
        ]

        if not identity:

            row[
                "validation_status"
            ] = "INVALID"

            row[
                "candidate_action"
            ] = "INVALID"

            row[
                "validation_errors"
            ].append(
                "scheme_name_missing"
            )

            continue

        warnings = []

        if identity in (
            cross_sheet_duplicates
        ):

            warnings.append(
                "cross_sheet_duplicate_name"
            )

        if identity in (
            existing_titles
        ):

            warnings.append(
                "existing_service_title_match"
            )

        if warnings:

            row[
                "validation_status"
            ] = "WARNING"

            row[
                "candidate_action"
            ] = "MERGE_REVIEW"

            row[
                "validation_warnings"
            ].extend(
                warnings
            )

        else:

            row[
                "validation_status"
            ] = "VALID"

            row[
                "candidate_action"
            ] = "CREATE"

    return {
        "duplicate_groups": len(
            cross_sheet_duplicates
        ),
        "duplicate_occurrences": (
            duplicate_occurrences
        ),
    }


def stage_workbook(
    workbook_path,
    expected_sha256,
    imported_by=None,
):

    workbook_path = Path(
        workbook_path
    )

    actual_sha256 = file_sha256(
        workbook_path
    )

    if (
        actual_sha256
        != expected_sha256
    ):
        raise ValueError(
            "Workbook SHA-256 does not "
            "match approved source."
        )

    for previous in (
        ImportBatch.objects
        .filter(
            file_sha256=actual_sha256
        )
        .order_by("-created_at")
    ):

        metadata = (
            previous.metadata
            if isinstance(
                previous.metadata,
                dict,
            )
            else {}
        )

        if (
            metadata.get(
                "operation"
            )
            == "workbook_staging"
        ):

            raise ValueError(
                "This exact workbook has "
                "already been staged as "
                f"ImportBatch #{previous.pk}. "
                "No duplicate staging run "
                "was created."
            )

    (
        rows,
        sheet_counts,
        family_counts,
    ) = extract_workbook_rows(
        workbook_path
    )

    duplicate_summary = (
        classify_scheme_candidates(
            rows
        )
    )

    structured_rows = (
        family_counts[
            "SCHEME_TABLE"
        ]
    )

    total_rows = len(
        rows
    )

    # --------------------------------------------------------
    # HARD SAFETY GATES
    # --------------------------------------------------------

    if (
        structured_rows
        != EXPECTED_STRUCTURED_ROWS
    ):
        raise ValueError(
            "Structured-row safety gate failed: "
            f"{structured_rows} != "
            f"{EXPECTED_STRUCTURED_ROWS}"
        )

    if (
        total_rows
        != EXPECTED_TOTAL_STAGED_ROWS
    ):
        raise ValueError(
            "Total staged-row safety gate failed: "
            f"{total_rows} != "
            f"{EXPECTED_TOTAL_STAGED_ROWS}"
        )

    if (
        duplicate_summary[
            "duplicate_groups"
        ]
        != EXPECTED_DUPLICATE_GROUPS
    ):
        raise ValueError(
            "Duplicate-group safety gate failed."
        )

    if (
        duplicate_summary[
            "duplicate_occurrences"
        ]
        != EXPECTED_DUPLICATE_OCCURRENCES
    ):
        raise ValueError(
            "Duplicate-occurrence safety "
            "gate failed."
        )

    validation_counts = Counter(
        row[
            "validation_status"
        ]
        for row in rows
    )

    action_counts = Counter(
        row[
            "candidate_action"
        ]
        for row in rows
    )

    metadata = {
        "operation": (
            "workbook_staging"
        ),
        "preview_only": True,
        "service_writes": False,
        "original_file_retained": False,
        "staged_normalized_rows_retained": (
            True
        ),
        "workbook_sha256_verified": True,
        "structured_scheme_rows": (
            structured_rows
        ),
        "staged_row_count": (
            total_rows
        ),
        "duplicate_name_groups": (
            duplicate_summary[
                "duplicate_groups"
            ]
        ),
        "duplicate_name_occurrences": (
            duplicate_summary[
                "duplicate_occurrences"
            ]
        ),
        "family_counts": dict(
            family_counts
        ),
        "validation_counts": dict(
            validation_counts
        ),
        "candidate_action_counts": dict(
            action_counts
        ),
        "sheets": [
            {
                "name": name,
                "staged_rows": (
                    sheet_counts[name]
                ),
            }
            for name
            in sorted(
                sheet_counts
            )
        ],
    }

    # --------------------------------------------------------
    # ONE TRANSACTION.
    # Batch and all rows commit together or not at all.
    # --------------------------------------------------------

    with transaction.atomic():

        batch = (
            ImportBatch.objects.create(
                source_type="XLSX",
                source_name=(
                    workbook_path.name
                ),
                source_identifier="",
                file_sha256=(
                    actual_sha256
                ),
                sheet_count=(
                    EXPECTED_SHEET_COUNT
                ),
                row_count=(
                    total_rows
                ),
                status="PREVIEWED",
                imported_by=(
                    imported_by
                ),
                metadata=metadata,
            )
        )

        objects = [
            ImportRow(
                import_batch=batch,
                sheet_name=(
                    row["sheet_name"]
                ),
                source_row_number=(
                    row[
                        "source_row_number"
                    ]
                ),
                source_key=(
                    row["source_key"]
                ),
                row_hash=(
                    row["row_hash"]
                ),
                raw_data=(
                    row["raw_data"]
                ),
                validation_status=(
                    row[
                        "validation_status"
                    ]
                ),
                validation_errors=(
                    row[
                        "validation_errors"
                    ]
                ),
                validation_warnings=(
                    row[
                        "validation_warnings"
                    ]
                ),
                candidate_action=(
                    row[
                        "candidate_action"
                    ]
                ),
            )
            for row in rows
        ]

        ImportRow.objects.bulk_create(
            objects,
            batch_size=200,
        )

        actual_created = (
            ImportRow.objects.filter(
                import_batch=batch
            ).count()
        )

        if (
            actual_created
            != total_rows
        ):
            raise RuntimeError(
                "Staging row-count verification "
                "failed inside transaction."
            )

    return (
        batch,
        {
            "sheet_counts": dict(
                sheet_counts
            ),
            "family_counts": dict(
                family_counts
            ),
            "validation_counts": dict(
                validation_counts
            ),
            "action_counts": dict(
                action_counts
            ),
            "duplicate_groups": (
                duplicate_summary[
                    "duplicate_groups"
                ]
            ),
            "duplicate_occurrences": (
                duplicate_summary[
                    "duplicate_occurrences"
                ]
            ),
            "total_rows": total_rows,
            "structured_rows": (
                structured_rows
            ),
        },
    )
