import csv
import hashlib
import io
import zipfile
from pathlib import Path, PurePosixPath

from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from openpyxl import load_workbook

from accounts.activity import log_activity

from .import_forms import ToolkitImportUploadForm
from .models import ImportBatch


MAX_PREVIEW_ROWS = 8
MAX_PREVIEW_COLUMNS = 30

MAX_XLSX_UNCOMPRESSED_BYTES = (
    250 * 1024 * 1024
)

MAX_XLSX_COMPRESSION_RATIO = 200


def can_import_toolkit(user):
    return (
        user.is_authenticated
        and (
            user.is_superuser
            or user.role in {
                "SUPER_ADMIN",
                "DATA_ADMIN",
            }
        )
    )


def normalize_value(value):
    if value is None:
        return ""

    return str(value).strip()


# =========================================================
# FILE FINGERPRINT
# =========================================================

def calculate_sha256(uploaded_file):
    digest = hashlib.sha256()

    uploaded_file.seek(0)

    for chunk in uploaded_file.chunks():
        digest.update(chunk)

    uploaded_file.seek(0)

    return digest.hexdigest()


# =========================================================
# CLIENT-REPORTED FILE MODIFICATION DATE
#
# This comes from browser File.lastModified.
# It is useful operational metadata but is not treated as
# a cryptographically trusted timestamp.
# =========================================================

def parse_source_modified_at(value):
    if not value:
        return None

    parsed = parse_datetime(value)

    if parsed is None:
        return None

    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed)

    return parsed


# =========================================================
# XLSX SECURITY VALIDATION
# =========================================================

def validate_xlsx_structure(uploaded_file):
    uploaded_file.seek(0)

    if not zipfile.is_zipfile(uploaded_file):
        uploaded_file.seek(0)

        raise ValueError(
            "The Excel file is not a valid XLSX archive."
        )

    uploaded_file.seek(0)

    with zipfile.ZipFile(uploaded_file) as archive:
        infos = archive.infolist()

        names = {
            info.filename
            for info in infos
        }

        required_files = {
            "[Content_Types].xml",
            "xl/workbook.xml",
        }

        if not required_files.issubset(names):
            raise ValueError(
                "The file does not contain a valid XLSX workbook."
            )

        # XLSX only. Macro-enabled content is not accepted.
        if any(
            name.lower().endswith(
                "vbaproject.bin"
            )
            for name in names
        ):
            raise ValueError(
                "Macro-enabled Excel files are not permitted."
            )

        total_compressed = 0
        total_uncompressed = 0

        for info in infos:
            path = PurePosixPath(
                info.filename
            )

            if (
                info.filename.startswith("/")
                or ".." in path.parts
            ):
                raise ValueError(
                    "Unsafe XLSX archive structure detected."
                )

            if info.flag_bits & 0x1:
                raise ValueError(
                    "Encrypted XLSX archives are not supported."
                )

            total_compressed += (
                info.compress_size
            )

            total_uncompressed += (
                info.file_size
            )

        if (
            total_uncompressed
            > MAX_XLSX_UNCOMPRESSED_BYTES
        ):
            raise ValueError(
                "The expanded Excel workbook is too large."
            )

        if (
            total_compressed > 0
            and (
                total_uncompressed
                / total_compressed
            )
            > MAX_XLSX_COMPRESSION_RATIO
        ):
            raise ValueError(
                "Unsafe Excel compression ratio detected."
            )

    uploaded_file.seek(0)


# =========================================================
# CSV SECURITY VALIDATION + PREVIEW
# =========================================================

def preview_csv(uploaded_file):
    uploaded_file.seek(0)

    raw = uploaded_file.read()

    uploaded_file.seek(0)

    if b"\x00" in raw:
        raise ValueError(
            "Binary content detected in CSV file."
        )

    text = None

    for encoding in (
        "utf-8-sig",
        "utf-8",
        "cp1252",
    ):
        try:
            text = raw.decode(encoding)
            break

        except UnicodeDecodeError:
            continue

    if text is None:
        raise ValueError(
            "CSV encoding could not be read."
        )

    csv.field_size_limit(
        1_000_000
    )

    stream = io.StringIO(text)

    reader = csv.reader(stream)

    first_row = next(
        reader,
        None
    )

    if first_row is None:
        return {
            "sheets": [
                {
                    "name": "CSV",
                    "headers": [],
                    "preview_rows": [],
                    "total_rows": 0,
                    "total_columns": 0,
                }
            ]
        }

    headers = [
        normalize_value(value)
        for value
        in first_row[
            :MAX_PREVIEW_COLUMNS
        ]
    ]

    preview_rows = []

    total_rows = 0
    total_columns = len(first_row)

    for row in reader:
        total_rows += 1

        total_columns = max(
            total_columns,
            len(row)
        )

        if (
            len(preview_rows)
            < MAX_PREVIEW_ROWS
        ):
            preview_rows.append(
                [
                    normalize_value(value)
                    for value
                    in row[
                        :MAX_PREVIEW_COLUMNS
                    ]
                ]
            )

    return {
        "sheets": [
            {
                "name": "CSV",
                "headers": headers,
                "preview_rows": preview_rows,
                "total_rows": total_rows,
                "total_columns": total_columns,
            }
        ]
    }


# =========================================================
# XLSX PREVIEW
# =========================================================

def preview_excel(uploaded_file):
    validate_xlsx_structure(
        uploaded_file
    )

    uploaded_file.seek(0)

    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    sheets = []

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(
                values_only=True
            )

            first_row = next(
                rows,
                None
            )

            if first_row is None:
                sheets.append(
                    {
                        "name": worksheet.title,
                        "headers": [],
                        "preview_rows": [],
                        "total_rows": 0,
                        "total_columns": 0,
                    }
                )

                continue

            headers = [
                normalize_value(value)
                for value
                in first_row[
                    :MAX_PREVIEW_COLUMNS
                ]
            ]

            preview_rows = []

            total_data_rows = 0

            for row in rows:
                total_data_rows += 1

                if (
                    len(preview_rows)
                    < MAX_PREVIEW_ROWS
                ):
                    preview_rows.append(
                        [
                            normalize_value(value)
                            for value
                            in row[
                                :MAX_PREVIEW_COLUMNS
                            ]
                        ]
                    )

            sheets.append(
                {
                    "name": worksheet.title,
                    "headers": headers,
                    "preview_rows": preview_rows,
                    "total_rows": total_data_rows,
                    "total_columns": (
                        worksheet.max_column
                        or 0
                    ),
                }
            )

    finally:
        workbook.close()

        uploaded_file.seek(0)

    return {
        "sheets": sheets
    }


# =========================================================
# IMPORT CENTER / SOURCE PREVIEW
# =========================================================

@user_passes_test(can_import_toolkit)
def import_center(request):
    preview = None
    import_error = None
    uploaded_name = None

    current_batch = None
    duplicate_batch = None

    if request.method == "POST":
        form = ToolkitImportUploadForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():
            uploaded_file = (
                form.cleaned_data["file"]
            )

            uploaded_name = (
                uploaded_file.name
            )

            extension = (
                Path(uploaded_file.name)
                .suffix
                .lower()
            )

            source_type = (
                "CSV"
                if extension == ".csv"
                else "XLSX"
            )

            try:
                file_sha256 = calculate_sha256(
                    uploaded_file
                )

                # Identical content, even if renamed.
                duplicate_batch = (
                    ImportBatch.objects
                    .filter(
                        file_sha256=file_sha256
                    )
                    .order_by(
                        "-created_at"
                    )
                    .first()
                )

                # Compare current version against the most
                # recent source with the same filename.
                previous_same_name = (
                    ImportBatch.objects
                    .filter(
                        source_type=source_type,
                        source_name=uploaded_name,
                    )
                    .order_by(
                        "-created_at"
                    )
                    .first()
                )

                if extension == ".csv":
                    preview = preview_csv(
                        uploaded_file
                    )

                elif extension == ".xlsx":
                    preview = preview_excel(
                        uploaded_file
                    )

                else:
                    raise ValueError(
                        "Unsupported file format."
                    )

                sheets = preview.get(
                    "sheets",
                    []
                )

                total_rows = sum(
                    sheet.get(
                        "total_rows",
                        0
                    )
                    for sheet in sheets
                )

                source_modified_at = (
                    parse_source_modified_at(
                        request.POST.get(
                            "source_modified_at"
                        )
                    )
                )

                metadata = {
                    "operation": "preview",
                    "preview_only": True,

                    "sheet_names": [
                        sheet.get(
                            "name",
                            ""
                        )
                        for sheet in sheets
                    ],

                    "sheets": [
                        {
                            "name": sheet.get(
                                "name",
                                ""
                            ),
                            "rows": sheet.get(
                                "total_rows",
                                0
                            ),
                            "columns": sheet.get(
                                "total_columns",
                                0
                            ),
                        }
                        for sheet in sheets
                    ],

                    "duplicate_detected": (
                        duplicate_batch
                        is not None
                    ),

                    "duplicate_of_id": (
                        duplicate_batch.pk
                        if duplicate_batch
                        else None
                    ),

                    "previous_same_name_id": (
                        previous_same_name.pk
                        if previous_same_name
                        else None
                    ),

                    "source_changed_since_previous": (
                        bool(
                            previous_same_name
                            and (
                                previous_same_name.file_sha256
                                != file_sha256
                            )
                        )
                    ),
                }

                current_batch = (
                    ImportBatch.objects.create(
                        source_type=source_type,

                        source_name=(
                            uploaded_name
                        ),

                        source_modified_at=(
                            source_modified_at
                        ),

                        file_sha256=file_sha256,

                        sheet_count=len(
                            sheets
                        ),

                        row_count=total_rows,

                        status="PREVIEWED",

                        imported_by=(
                            request.user
                        ),

                        metadata=metadata,
                    )
                )

                # Do NOT put confidential filenames,
                # row contents or cell contents into the
                # generic ActivityLog.
                log_activity(
                    request,
                    "IMPORT",
                    "Previewed toolkit source file.",
                    target_type=(
                        "import_batch"
                    ),
                    target_id=(
                        current_batch.pk
                    ),
                    metadata={
                        "operation": (
                            "preview"
                        ),
                        "batch_id": (
                            current_batch.pk
                        ),
                        "source_type": (
                            source_type
                        ),
                        "sheet_count": (
                            len(sheets)
                        ),
                        "row_count": (
                            total_rows
                        ),
                        "duplicate_detected": (
                            duplicate_batch
                            is not None
                        ),
                    },
                )

            except Exception as exc:
                import_error = str(
                    exc
                ) or (
                    "The source file could "
                    "not be read."
                )

    else:
        form = ToolkitImportUploadForm()

    return render(
        request,
        "toolkit/admin/import_center.html",
        {
            "form": form,
            "preview": preview,
            "import_error": import_error,
            "uploaded_name": uploaded_name,
            "current_batch": current_batch,
            "duplicate_batch": duplicate_batch,
        }
    )


# =========================================================
# SOURCE / IMPORT HISTORY
# =========================================================

@user_passes_test(can_import_toolkit)
def import_history(request):
    source_filter = request.GET.get(
        "source",
        ""
    ).strip()

    status_filter = request.GET.get(
        "status",
        ""
    ).strip()

    query = request.GET.get(
        "q",
        ""
    ).strip()

    batches = (
        ImportBatch.objects
        .select_related(
            "imported_by"
        )
        .all()
    )

    valid_sources = {
        value
        for value, label
        in ImportBatch.SOURCE_CHOICES
    }

    valid_statuses = {
        value
        for value, label
        in ImportBatch.STATUS_CHOICES
    }

    if source_filter in valid_sources:
        batches = batches.filter(
            source_type=source_filter
        )

    if status_filter in valid_statuses:
        batches = batches.filter(
            status=status_filter
        )

    if query:
        batches = batches.filter(
            Q(
                source_name__icontains=query
            )
            | Q(
                file_sha256__icontains=query
            )
        )

    all_batches = ImportBatch.objects.all()

    total_rows_seen = (
        all_batches.aggregate(
            total=Sum("row_count")
        )["total"]
        or 0
    )

    duplicate_count = sum(
        1
        for metadata
        in all_batches.values_list(
            "metadata",
            flat=True
        )
        if (
            isinstance(
                metadata,
                dict
            )
            and metadata.get(
                "duplicate_detected"
            )
        )
    )

    context = {
        "batches": batches[:250],

        "source_filter": source_filter,
        "status_filter": status_filter,
        "query": query,

        "source_choices": (
            ImportBatch.SOURCE_CHOICES
        ),

        "status_choices": (
            ImportBatch.STATUS_CHOICES
        ),

        "total_batches": (
            all_batches.count()
        ),

        "total_rows_seen": (
            total_rows_seen
        ),

        "duplicate_count": (
            duplicate_count
        ),
    }

    return render(
        request,
        "toolkit/admin/import_history.html",
        context
    )


# =========================================================
# SOURCE HISTORY DETAIL
# =========================================================

@user_passes_test(can_import_toolkit)
def import_history_detail(
    request,
    batch_id
):
    batch = get_object_or_404(
        ImportBatch.objects
        .select_related(
            "imported_by"
        ),
        pk=batch_id
    )

    duplicate_of = None
    previous_same_name = None

    if isinstance(
        batch.metadata,
        dict
    ):
        duplicate_id = (
            batch.metadata.get(
                "duplicate_of_id"
            )
        )

        previous_id = (
            batch.metadata.get(
                "previous_same_name_id"
            )
        )

        if duplicate_id:
            duplicate_of = (
                ImportBatch.objects
                .filter(
                    pk=duplicate_id
                )
                .first()
            )

        if previous_id:
            previous_same_name = (
                ImportBatch.objects
                .filter(
                    pk=previous_id
                )
                .first()
            )

    return render(
        request,
        "toolkit/admin/import_history_detail.html",
        {
            "batch": batch,
            "duplicate_of": duplicate_of,
            "previous_same_name": (
                previous_same_name
            ),
        }
    )
